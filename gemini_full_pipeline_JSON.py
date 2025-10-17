"""
Full Gemini pipeline: boundary detection → compression → LLM analysis (JSON output)
Processes all sheets in a workbook end-to-end.
"""
import json
import time
import sys
import argparse
import re
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Tuple, Optional, List

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("Note: tqdm not installed. Install with 'pip install tqdm' for progress bars.")

try:
    import vertexai
    from vertexai.generative_models import GenerativeModel
    VERTEXAI_AVAILABLE = True
except ImportError:
    VERTEXAI_AVAILABLE = False
    print("ERROR: google-cloud-aiplatform not installed.")
    print("Install with: pip install google-cloud-aiplatform")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Sheet ordering may be alphabetical.")
    print("Install with: pip install openpyxl")

# Import boundary detection and compression modules
from boundary_detector.core.batch_detector import BatchBoundaryDetector
from compression.format_compression import compress_formats
from compression.inverted_index_compression import compress_format_result_json, save_compressed_json

# Base paths
BASE_DIR = Path(__file__).parent
SERVICE_ACCOUNT_FILE = BASE_DIR / "gemini_service_account.json"

# Gemini prompt with label/data separation and multiple time series detection - JSON OUTPUT
ANALYSIS_PROMPT = """You are an expert financial analyst tasked with creating a high-level summary of an Excel spreadsheet's structure. Your goal is to identify and describe the key logical data sections (e.g., "Income Statement," "Balance Sheet," "KPI Dashboard") without listing every single row or label. The output must be a structured JSON document that enables retrieval of specific, relevant sections by cell range.

# CORE OBJECTIVE
Identify major financial or data sections within this Excel sheet, describe their purpose, and define their precise boundaries and time series context.

# CRITICAL: MULTIPLE TIME SERIES DETECTION
Many spreadsheets have MULTIPLE time series sections side-by-side:
- **Annual labels** (e.g., rows or columns showing years like 2015, 2016, 2017...)
- **Monthly/Quarterly labels** (e.g., rows or columns showing monthly dates like 2015-01-31, 2015-02-28... or quarters like Q1 2024, Q2 2024...)

When analyzing time series:
1. Look for MULTIPLE date header rows or columns with different frequencies
2. Check if there are gaps between column groups (suggesting separate time series)
3. Identify if data spans different date column groups
4. Report ALL time series horizons found, not just one
5. **IMPORTANT**: If there is only one time series range on a whole sheet or for a large section of a sheet, that series often applies as a column or row header across multiple sections.

# OUTPUT REQUIREMENTS
Generate a JSON document with the following structure:

{
  "sheet_name": "The name of the analyzed sheet",
  "sections": [
    {
      "section_name": "Name of the section (e.g., 'Income Statement')",
      "section_type": "Classification (e.g., 'Standard P&L', 'Balance Sheet', 'Key Metrics Table')",
      "description": "Brief explanation of what this section represents and its business purpose",
      "cell_range": {
        "row_headers_location": "The exact and complete cell range (e.g., 'A5:FS88')",
        "column_headers_location": "The exact and complete cell range (e.g., 'A5:FS88')",
        "data_range": "The exact and complete cell range (e.g., 'A5:FS88')"
      },
      "time_series": [
        {
          "label": "Description (e.g., 'Annual', 'Monthly')",
          "start_date": "Start date/period (e.g., '2015', 'Year 1', '1Q24', etc.)",
          "end_date": "End date/period (e.g., '2025', 'Year 10', '4Q26', etc.)",
          "frequency": "Time interval (e.g., 'Annual', 'Monthly', 'Quarterly')",
          "cell_range": "Cell range where this time series appears",
          "anchor_points": "Optional: Key date anchor points if provided in date_series_definitions"
        }
      ],
      "key_components": ["List of important labels/metrics"],
      "notes": "Any deviations from standard reports or special calculations"
    }
  ]
}

# GUIDING PRINCIPLES
- **High-Level, Not Granular**: Focus on the forest, not the trees. Summarize sections, don't transcribe them.
- **Precision is Key**: Cell ranges and date locations must be exact for future retrieval.
- **Context Over Content**: Explain the *purpose* of a section, not just what's in it.
- **Multiple Time Series**: If you see evidence of multiple date columns with different frequencies, report them ALL.
- **Date Anchor Points**: If the JSON contains a `date_series_definitions` section with `anchor_points`, include these in your Time Series Details to help locate specific dates within long series.

# IMPORTANT
Your response must be ONLY valid JSON. Do not include markdown code fences, explanatory text, or any content outside the JSON structure.

Now, analyze the provided compressed JSON and generate the complete, high-level structured JSON documentation."""


def find_excel_files(raw_data_dir: Path) -> List[Path]:
    """Find all Excel files in a directory"""
    excel_files = []
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(raw_data_dir.glob(ext))
    return sorted(excel_files)


def get_dataset_paths(dataset_name: str) -> Tuple[Path, Path, Path]:
    """
    Get paths for a dataset

    Returns:
        Tuple of (raw_data_dir, processing_outputs_dir, dataset_base_dir)
    """
    dataset_base = BASE_DIR / "datasets" / dataset_name
    raw_data_dir = dataset_base / "raw_data"
    processing_outputs_dir = dataset_base / "processing_outputs"

    return raw_data_dir, processing_outputs_dir, dataset_base


def select_excel_file(excel_files: List[Path]) -> Optional[Path]:
    """
    Prompt user to select an Excel file if multiple are found

    Returns:
        Selected file path or None if cancelled
    """
    if len(excel_files) == 0:
        return None

    if len(excel_files) == 1:
        return excel_files[0]

    # Multiple files - prompt user
    print("\nMultiple Excel files found:")
    for i, file_path in enumerate(excel_files, 1):
        print(f"  {i}. {file_path.name}")

    while True:
        try:
            choice = input(f"\nSelect a file (1-{len(excel_files)}) or 'q' to quit: ").strip()
            if choice.lower() == 'q':
                return None

            idx = int(choice) - 1
            if 0 <= idx < len(excel_files):
                return excel_files[idx]
            else:
                print(f"Invalid choice. Please enter a number between 1 and {len(excel_files)}.")
        except ValueError:
            print("Invalid input. Please enter a number or 'q' to quit.")
        except (EOFError, KeyboardInterrupt):
            print("\nCancelled.")
            return None


def sanitize_filename(name: str) -> str:
    """Convert sheet name to safe filename"""
    safe = name.replace('/', '_').replace('\\', '_').replace(':', '_')
    safe = safe.replace('*', '_').replace('?', '_').replace('"', '_')
    safe = safe.replace('<', '_').replace('>', '_').replace('|', '_')
    return safe


def get_workbook_sheet_order(workbook_path: Path) -> List[str]:
    """
    Get the sheet names in the order they appear in the workbook.

    Returns:
        List of sheet names in workbook order
    """
    if not OPENPYXL_AVAILABLE:
        return []

    try:
        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"WARNING: Could not load workbook to determine sheet order: {e}")
        print("Falling back to alphabetical sorting")
        return []


def process_sheet(
    sheet_name: str,
    workbook_path: Path,
    boundary_result,
    project_id: str,
    json_compressed_dir: Path,
    json_analysis_dir: Path,
    max_retries: int = 5
) -> Tuple[str, bool, Optional[str]]:
    """
    Process a single sheet: compress to JSON → analyze with Gemini → output JSON

    Args:
        sheet_name: Name of the sheet to process
        workbook_path: Path to Excel workbook
        boundary_result: BoundaryResult object from boundary detection
        project_id: Google Cloud project ID
        json_compressed_dir: Directory to save compressed JSON output
        json_analysis_dir: Directory to save JSON analysis output
        max_retries: Max retry attempts for rate limits

    Returns:
        Tuple of (sheet_name, success, error_message)
    """
    if not VERTEXAI_AVAILABLE:
        return (sheet_name, False, "vertexai SDK not installed")

    safe_name = sanitize_filename(sheet_name)
    json_compressed_file = json_compressed_dir / f"{safe_name}_compressed.json"
    json_analysis_file = json_analysis_dir / f"{safe_name}_analysis.json"

    # Skip if JSON analysis already exists
    if json_analysis_file.exists():
        return (sheet_name, True, None)

    try:
        # Step 1: Compress sheet to JSON (if not already done)
        if not json_compressed_file.exists():
            # Compress formats using the boundary result
            format_result = compress_formats(
                file_path=str(workbook_path),
                sheet_name=sheet_name,
                boundary_result=boundary_result
            )

            # Convert FormatResult to JSON
            json_output = format_result.to_json()

            # Apply inverted index compression
            compressed_json = compress_format_result_json(json_output)

            # Save compressed JSON
            json_compressed_dir.mkdir(parents=True, exist_ok=True)
            save_compressed_json(compressed_json, str(json_compressed_file))

        # Step 2: Load compressed JSON
        with open(json_compressed_file, 'r', encoding='utf-8') as f:
            compressed_json = json.load(f)

        # Step 3: Analyze with Gemini (with retry logic)
        for attempt in range(max_retries):
            try:
                # Convert JSON to string for prompt
                json_str = json.dumps(compressed_json, indent=2)

                # Create the full prompt
                full_prompt = f"{ANALYSIS_PROMPT}\n\n# Compressed JSON Data\n\n```json\n{json_str}\n```"

                # Initialize Vertex AI
                vertexai.init(project=project_id, location="us-central1")

                # Create model
                model = GenerativeModel("gemini-2.0-flash-exp")

                # Add a small delay to avoid overwhelming the API (spread out requests)
                if attempt > 0:
                    time.sleep(2)  # Wait 2s before retry

                # Generate response
                response = model.generate_content(
                    full_prompt,
                    generation_config={
                        "max_output_tokens": 8192,
                        "temperature": 0.3,
                    }
                )

                # Extract text
                analysis = response.text

                if not analysis or len(analysis.strip()) == 0:
                    return (sheet_name, False, "Empty response from Gemini")

                # Clean up response - remove markdown code fences if present
                analysis = analysis.strip()
                if analysis.startswith('```json'):
                    analysis = analysis[7:]  # Remove ```json
                if analysis.startswith('```'):
                    analysis = analysis[3:]  # Remove ```
                if analysis.endswith('```'):
                    analysis = analysis[:-3]  # Remove trailing ```
                analysis = analysis.strip()

                # Validate it's valid JSON
                try:
                    parsed_json = json.loads(analysis)
                except json.JSONDecodeError as e:
                    return (sheet_name, False, f"Invalid JSON from Gemini: {str(e)[:100]}")

                # Save JSON analysis
                json_analysis_dir.mkdir(parents=True, exist_ok=True)
                with open(json_analysis_file, 'w', encoding='utf-8') as f:
                    json.dump(parsed_json, f, indent=2)

                return (sheet_name, True, None)

            except Exception as e:
                error_msg = str(e)

                # Check if it's a rate limit error (multiple possible error messages)
                is_rate_limit = (
                    "429" in error_msg or
                    "quota" in error_msg.lower() or
                    "rate" in error_msg.lower() or
                    "resource_exhausted" in error_msg.lower() or
                    "RESOURCE_EXHAUSTED" in error_msg
                )

                if is_rate_limit:
                    if attempt < max_retries - 1:
                        # Exponential backoff: 10s, 20s, 40s, 80s
                        wait_time = (2 ** attempt) * 10
                        print(f"\n  ⏸️  Rate limit hit for '{sheet_name}', waiting {wait_time}s... (attempt {attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        # Max retries exceeded for rate limit
                        return (sheet_name, False, f"Rate limit - max retries exceeded: {error_msg[:100]}")

                # Non-rate-limit error - fail immediately
                return (sheet_name, False, error_msg[:200])

        return (sheet_name, False, "Max retries exceeded")

    except Exception as e:
        return (sheet_name, False, str(e))


def main():
    """Run full Gemini pipeline on entire workbook with JSON output"""

    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="Run Gemini pipeline on a dataset (JSON output)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python gemini_full_pipeline_JSON.py alphasense
  python gemini_full_pipeline_JSON.py evergreen-asset-acquisition
  python gemini_full_pipeline_JSON.py pendo-sales-quota-build
        """
    )
    parser.add_argument(
        "dataset",
        help="Name of the dataset folder (e.g., 'alphasense', 'evergreen-asset-acquisition')"
    )
    args = parser.parse_args()

    dataset_name = args.dataset

    # Get dataset paths
    raw_data_dir, processing_outputs_dir, dataset_base = get_dataset_paths(dataset_name)

    # Verify dataset exists
    if not dataset_base.exists():
        print(f"ERROR: Dataset folder not found: {dataset_base}")
        print(f"\nAvailable datasets:")
        datasets_dir = BASE_DIR / "datasets"
        if datasets_dir.exists():
            for item in sorted(datasets_dir.iterdir()):
                if item.is_dir():
                    print(f"  - {item.name}")
        return

    if not raw_data_dir.exists():
        print(f"ERROR: raw_data folder not found: {raw_data_dir}")
        return

    # Find Excel files
    excel_files = find_excel_files(raw_data_dir)
    if not excel_files:
        print(f"ERROR: No Excel files found in {raw_data_dir}")
        return

    # Select Excel file
    workbook_path = select_excel_file(excel_files)
    if workbook_path is None:
        print("No file selected. Exiting.")
        return

    # Setup output directories
    json_compressed_dir = processing_outputs_dir / "json"
    json_analysis_dir = processing_outputs_dir / "json_analysis"

    print(f"\n{'='*80}")
    print(f"GEMINI FULL PIPELINE (JSON OUTPUT)")
    print(f"{'='*80}")
    print(f"Dataset: {dataset_name}")
    print(f"Workbook: {workbook_path}")
    print(f"JSON Compressed Output: {json_compressed_dir}")
    print(f"JSON Analysis Output: {json_analysis_dir}")
    print(f"{'='*80}\n")

    if not VERTEXAI_AVAILABLE:
        print("\nERROR: Cannot proceed without google-cloud-aiplatform")
        print("Install with: pip install google-cloud-aiplatform")
        return

    # Set service account environment variable
    import os
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(SERVICE_ACCOUNT_FILE)

    # Load project ID from service account
    if not SERVICE_ACCOUNT_FILE.exists():
        print(f"ERROR: Service account file not found: {SERVICE_ACCOUNT_FILE}")
        return

    with open(SERVICE_ACCOUNT_FILE, 'r') as f:
        sa_data = json.load(f)
    project_id = sa_data['project_id']

    # Step 1: Detect boundaries for all sheets
    print("STEP 1: Detecting boundaries for all sheets...")
    start_time = time.time()

    detector = BatchBoundaryDetector()
    boundary_results = detector.detect_all_sheets(
        str(workbook_path),
        detect_charts=False,  # Disable chart detection for speed (saves ~30s)
        parallel=True,
        max_workers=5
    )

    elapsed = time.time() - start_time
    print(f"  Boundary detection complete: {elapsed:.1f}s")
    print(f"  Found {len(boundary_results)} sheets\n")

    # Step 2: Process all sheets (compression + LLM analysis) in parallel
    print("STEP 2: Compressing sheets and analyzing with Gemini (JSON output)...")
    print("  Using 3 parallel workers to respect API rate limits")
    start_time = time.time()
    success_count = 0
    error_count = 0
    errors = []

    # Use 3 workers to balance speed with API rate limits
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {
            executor.submit(
                process_sheet,
                sheet_name,
                workbook_path,
                boundary_result,
                project_id,
                json_compressed_dir,
                json_analysis_dir
            ): sheet_name
            for sheet_name, boundary_result in boundary_results.items()
        }

        # Use tqdm if available
        if TQDM_AVAILABLE:
            iterator = tqdm(as_completed(futures), total=len(futures), desc="Processing")
        else:
            iterator = as_completed(futures)

        # Collect results
        for future in iterator:
            sheet_name, success, error = future.result()

            if success:
                success_count += 1
                if not TQDM_AVAILABLE:
                    print(f"  OK {sheet_name}")
            else:
                error_count += 1
                errors.append((sheet_name, error))
                if not TQDM_AVAILABLE:
                    print(f"  FAIL {sheet_name}: {error}")

    elapsed = time.time() - start_time
    print(f"\n  Compression and analysis complete: {elapsed:.1f}s ({elapsed/60:.1f} minutes)")

    # Summary
    print(f"\n{'='*80}")
    print(f"PIPELINE COMPLETE")
    print(f"{'='*80}")
    print(f"Total sheets: {len(boundary_results)}")
    print(f"Successful: {success_count}/{len(boundary_results)}")
    print(f"Failed: {error_count}/{len(boundary_results)}")
    if errors:
        print(f"\nFailed sheets:")
        for sheet_name, error in errors:
            print(f"  - {sheet_name}: {error}")
    print(f"\nProcessing time: {elapsed:.1f}s ({elapsed/60:.1f} minutes)")
    print(f"\nOutputs saved to:")
    print(f"  JSON Compressed: {json_compressed_dir}")
    print(f"  JSON Analysis: {json_analysis_dir}")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()
