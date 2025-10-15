"""
Combine all sheet markdown analyses into a single master document.
Creates a table of contents and clearly differentiates between sheets.
Preserves the original workbook sheet order.
"""
from pathlib import Path
import re
from typing import List, Dict, Tuple
from openpyxl import load_workbook

# Paths
BASE_DIR = Path(__file__).parent
WORKBOOK_PATH = BASE_DIR / "sample_data" / "sample.xlsx"
MARKDOWN_DIR = BASE_DIR / "output" / "batch" / "markdown"
OUTPUT_FILE = BASE_DIR / "output" / "batch" / "MASTER_WORKBOOK_ANALYSIS.md"


def get_workbook_sheet_order(workbook_path: Path) -> List[str]:
    """
    Get the sheet names in the order they appear in the workbook.

    Returns:
        List of sheet names in workbook order
    """
    try:
        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"WARNING: Could not load workbook to determine sheet order: {e}")
        print("Falling back to alphabetical sorting")
        return []


def sanitize_filename(name: str) -> str:
    """Convert sheet name to safe filename (matches gemini_full_pipeline.py logic)"""
    safe = name.replace('/', '_').replace('\\', '_').replace(':', '_')
    safe = safe.replace('*', '_').replace('?', '_').replace('"', '_')
    safe = safe.replace('<', '_').replace('>', '_').replace('|', '_')
    return safe


def parse_markdown_file(file_path: Path) -> Dict:
    """
    Parse a markdown file to extract sheet name and sections.

    Returns:
        Dict with 'sheet_name', 'sections', and 'content'
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Remove markdown code fence if present
    content = re.sub(r'^```markdown\n', '', content)
    content = re.sub(r'\n```$', '', content)

    # Extract sheet name from "## 1. Spreadsheet Overview" section
    sheet_name_match = re.search(r'\*\*Sheet Name\*\*:\s*(.+?)(?:\n|$)', content)
    sheet_name = sheet_name_match.group(1).strip() if sheet_name_match else file_path.stem.replace('_analysis', '')

    # Extract all section headers (### lines)
    sections = []
    for match in re.finditer(r'^### (.+?)$', content, re.MULTILINE):
        section_title = match.group(1).strip()
        sections.append(section_title)

    return {
        'sheet_name': sheet_name,
        'file_path': file_path,
        'sections': sections,
        'content': content
    }


def create_master_toc(parsed_files: List[Dict]) -> str:
    """
    Create a master table of contents for all sheets.

    Returns:
        Markdown string with TOC
    """
    toc = ["# Master Workbook Analysis", ""]
    toc.append("This document provides a comprehensive analysis of all sheets in the Excel workbook.")
    toc.append("")
    toc.append("---")
    toc.append("")
    toc.append("## Table of Contents")
    toc.append("")

    for idx, parsed in enumerate(parsed_files, 1):
        sheet_name = parsed['sheet_name']
        # Create anchor link (lowercase, replace spaces with hyphens)
        anchor = sheet_name.lower().replace(' ', '-').replace('/', '').replace('\\', '').replace('-->', '').replace('>', '').replace('<', '').replace('|', '').replace(':', '')
        toc.append(f"{idx}. [{sheet_name}](#{anchor})")

        # Add subsections if any
        if parsed['sections']:
            for section in parsed['sections']:
                # Create section anchor
                section_anchor = section.lower().replace(' ', '-').replace('/', '').replace('\\', '').replace('(', '').replace(')', '').replace('.', '').replace(',', '').replace(':', '')
                toc.append(f"   - [{section}](#{anchor}-{section_anchor})")

    toc.append("")
    toc.append("---")
    toc.append("")

    return '\n'.join(toc)


def format_sheet_section(parsed: Dict, sheet_number: int) -> str:
    """
    Format a single sheet's content with clear differentiation.

    Returns:
        Markdown string for the sheet
    """
    sheet_name = parsed['sheet_name']
    content = parsed['content']

    # Create a clear sheet header
    output = []
    output.append("")
    output.append("=" * 100)
    output.append(f"# SHEET {sheet_number}: {sheet_name}")
    output.append("=" * 100)
    output.append("")

    # Add the sheet content (remove the first # if it exists to avoid nesting issues)
    content = re.sub(r'^# ', '## ', content, flags=re.MULTILINE)

    output.append(content)
    output.append("")

    return '\n'.join(output)


def combine_markdowns():
    """Main function to combine all markdown files."""

    print(f"\nCOMBINING MARKDOWN ANALYSES")
    print(f"=" * 80)
    print(f"Input directory: {MARKDOWN_DIR}")
    print(f"Output file: {OUTPUT_FILE}")
    print(f"=" * 80)
    print()

    # Get workbook sheet order
    print("Loading workbook to determine sheet order...")
    workbook_sheet_order = get_workbook_sheet_order(WORKBOOK_PATH)
    if workbook_sheet_order:
        print(f"Found {len(workbook_sheet_order)} sheets in workbook")
    print()

    # Find all markdown files and create a lookup by sheet name
    markdown_files_by_sheet = {}
    for file_path in MARKDOWN_DIR.glob("*_analysis.md"):
        # Extract original sheet name from filename
        safe_name = file_path.stem.replace('_analysis', '')
        # Try to match to original sheet name
        for sheet_name in workbook_sheet_order:
            if sanitize_filename(sheet_name) == safe_name:
                markdown_files_by_sheet[sheet_name] = file_path
                break
        else:
            # If no match found, store with sanitized name as fallback
            markdown_files_by_sheet[safe_name] = file_path

    if not markdown_files_by_sheet:
        print("ERROR: No markdown files found!")
        return

    print(f"Found {len(markdown_files_by_sheet)} markdown files")
    print()

    # Parse all files in workbook order
    print("Parsing markdown files...")
    parsed_files = []

    # First, add sheets in workbook order if we have the sheet order
    if workbook_sheet_order:
        for sheet_name in workbook_sheet_order:
            if sheet_name in markdown_files_by_sheet:
                file_path = markdown_files_by_sheet[sheet_name]
                try:
                    parsed = parse_markdown_file(file_path)
                    parsed_files.append(parsed)
                    print(f"  + {parsed['sheet_name']} ({len(parsed['sections'])} sections)")
                except Exception as e:
                    print(f"  x {file_path.name}: {e}")
    else:
        # Fallback to alphabetical if we couldn't get workbook order
        print("Using alphabetical order...")
        for sheet_name in sorted(markdown_files_by_sheet.keys()):
            file_path = markdown_files_by_sheet[sheet_name]
            try:
                parsed = parse_markdown_file(file_path)
                parsed_files.append(parsed)
                print(f"  + {parsed['sheet_name']} ({len(parsed['sections'])} sections)")
            except Exception as e:
                print(f"  x {file_path.name}: {e}")

    print()

    # Create master document
    print("Creating master document...")
    master_content = []

    # Add table of contents
    master_content.append(create_master_toc(parsed_files))

    # Add each sheet
    for idx, parsed in enumerate(parsed_files, 1):
        sheet_section = format_sheet_section(parsed, idx)
        master_content.append(sheet_section)

    # Write to file
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(master_content))

    print(f"SUCCESS: Master document created successfully!")
    print()
    print(f"Summary:")
    print(f"  - Total sheets: {len(parsed_files)}")
    print(f"  - Total sections: {sum(len(p['sections']) for p in parsed_files)}")
    print(f"  - Output file: {OUTPUT_FILE}")
    print(f"  - File size: {OUTPUT_FILE.stat().st_size:,} bytes")
    print()


if __name__ == "__main__":
    combine_markdowns()
