"""Test how accurate openpyxl's max_row and max_column are"""

import openpyxl

# Expected boundaries from manual verification
EXPECTED_BOUNDARIES = {
    'Summary': ('BW', 66),
    'Master Ctrl': ('E', 33),
    'Dash Control': ('FS', 66),
    'Dash': ('FS', 373),
    'CAC by Segment': ('FS', 141),
    'Payback Period': ('Q', 805),
    'Fin CTRL': ('FS', 189),
    'Fin': ('FS', 836),
    'Debt Compliance': ('ED', 53),
    'ARR and Revenue -->': (None, None),
    'ARR and Rev CTRL': ('FS', 318),
    'Sales Prod Input': ('Q', 31),
    'ARR and Revenue': ('FS', 233),
    'Sales CTRL': ('FS', 575),
    'Sales Role Input': ('DU', 59),
    'Sales Reps': ('FS', 524),
    'Clients CTRL': ('FS', 78),
    'Clients': ('FS', 68),
    'Retention CTRL': ('FS', 46),
    'Retention': ('FS', 86),
    'Renewed': ('N', 7899),
    'Renewal Base': ('N', 1858),
    'Deferred Build': ('FS', 34),
    'Expenses -->': (None, None),
    'Headcount and Salaries CTRL': ('FS', 152),
    'Headcount and Salaries': ('FS', 194),
    'Opex CTRL': ('FS', 2492),
    'Operating Expenses': ('FS', 231),
    'Content': ('FS', 399),
    'Engineering': ('FS', 64),
    'Executive': ('FS', 399),
    'Finance & Admin': ('FS', 403),
    'Marketing': ('FS', 385),
    'Product': ('FS', 398),
    'Sales': ('FS', 405),
    'Commission Waterfall': ('FS', 583),
    'Bad Debt': ('CT', 24),
    'COGS CTRL': ('FS', 1590),
    'Cost of Goods Sold': ('FS', 817),
    'Total FDS Cost': ('DD', 47),
    'FDS AMR': ('DD', 89),
    'Direct Broker': ('FS', 255),
    'Additional FDS Content': ('DF', 112),
    'FDS User Growth': ('DD', 297),
    'FDS RT Minimums': ('J', 26),
    'User Split Support-->': (None, None),
    'Product User Splits': ('BP', 119),
    'TR BRM': ('AB', 44),
    'AMR': ('AA', 25),
    'LN': ('T', 9),
    'DJ': ('T', 7),
    'TR Transcripts': ('T', 17),
    'FS Transcripts': ('T', 17),
    'Filings': ('T', 8),
    'Support>>>': (None, None),
    'Detailed Income Statement': ('BV', 167),
    'Detailed Balance Sheet': ('BV', 102),
    'Detailed Cash Flow Satement': ('BV', 86),
    'Key Metrics': ('CG', 499),
    'Revenue by Client': ('BU', 22),
    'Client Support': ('EE', 41),
    'Subscriber Support': ('EC', 27),
    'Headcount Support': ('CI', 77),
    'Salary Support': ('GD', 86),
    'Bonus Support': ('CX', 75),
    'Bonus Support (Sales non Ops)': ('CX', 54),
    'Bonus Support (Sales Ops Only)': ('CX', 75),
    'OpEx - Content': ('U', 138),
    'OpEx - Engineering': ('U', 138),
    'OpEx - Executive': ('U', 138),
    'OpEx - Finance & Admin': ('U', 138),
    'OpEx - Marketing': ('U', 138),
    'OpEx - Product': ('U', 138),
    'OpEx - Sales': ('U', 138),
    'Historical Quota& Productivity': ('BH', 26),
    'Fixed Asset Depreciation': ('DV', 22),
    'INTERNAL>>>': (None, None),
    'COGS vs. Marginal': ('E', 105),
    'Benchmarking': ('O', 80),
    'Accrued Expenses': ('AC', 49),
    'Prepaid Expenses': ('Z', 81),
    'Accrued Wages': ('AA', 14),
    'Cap R&D Exisiting Runoff': ('BJ', 19),
    'Deposits': ('Y', 3),
    'Tekes Loans': ('AZ', 61),
}

def column_letter_to_index(col_letter):
    """Convert column letter to 0-indexed number"""
    if col_letter is None:
        return None
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def column_index_to_letter(col_index):
    """Convert 0-indexed column number to letter"""
    if col_index is None:
        return None
    result = []
    col_index += 1
    while col_index > 0:
        col_index -= 1
        result.append(chr(col_index % 26 + ord('A')))
        col_index //= 26
    return ''.join(reversed(result))

# Load workbook
file_path = "sample_data/sample.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

print("Testing openpyxl max_row and max_column accuracy")
print("=" * 100)
print(f"{'Sheet Name':<40} {'Expected':<15} {'Openpyxl':<15} {'Column Diff':<12} {'Row Diff':<12} {'Status'}")
print("=" * 100)

total_sheets = 0
passed_sheets = 0
failed_sheets = 0

# Test only a representative subset of sheets for speed
TEST_SHEETS = [
    'Summary',           # Known issue: -1 row, -1 col
    'Engineering',       # Known issue: +332 rows
    'Renewed',          # Known issue: -300 rows
    'Master Ctrl',      # Simple case
    'Revenue by Client', # Corrected expected value
    'Dash',             # Large sheet
    'ARR and Revenue -->', # Empty sheet
]

for sheet_name in TEST_SHEETS:
    if sheet_name not in EXPECTED_BOUNDARIES:
        continue

    expected_col_letter, expected_row = EXPECTED_BOUNDARIES[sheet_name]
    total_sheets += 1

    sheet = wb[sheet_name]
    openpyxl_max_col = sheet.max_column
    openpyxl_max_row = sheet.max_row

    detected_col_letter = column_index_to_letter(openpyxl_max_col - 1)  # Convert to 0-indexed
    detected_row = openpyxl_max_row

    # Handle empty sheets
    if expected_col_letter is None:
        expected_display = "EMPTY"
        detected_display = f"{detected_col_letter}{detected_row}"

        if openpyxl_max_col <= 2 and openpyxl_max_row <= 2:
            status = "PASS"
            passed_sheets += 1
            col_diff = "N/A"
            row_diff = "N/A"
        else:
            status = "FAIL"
            failed_sheets += 1
            col_diff = f"+{openpyxl_max_col - 1}"
            row_diff = f"+{openpyxl_max_row - 1}"
    else:
        expected_display = f"{expected_col_letter}{expected_row}"
        detected_display = f"{detected_col_letter}{detected_row}"

        expected_col_idx = column_letter_to_index(expected_col_letter)
        col_diff = (openpyxl_max_col - 1) - expected_col_idx
        row_diff = openpyxl_max_row - expected_row

        if col_diff == 0 and row_diff == 0:
            status = "PASS"
            passed_sheets += 1
            col_diff_str = "0"
            row_diff_str = "0"
        else:
            status = "FAIL"
            failed_sheets += 1
            col_diff_str = f"{col_diff:+d}" if col_diff != 0 else "0"
            row_diff_str = f"{row_diff:+d}" if row_diff != 0 else "0"

        col_diff = col_diff_str
        row_diff = row_diff_str

    print(f"{sheet_name:<40} {expected_display:<15} {detected_display:<15} {col_diff:<12} {row_diff:<12} {status}")

wb.close()

print("=" * 100)
print(f"\nSummary:")
print(f"  Total sheets: {total_sheets}")
print(f"  Passed: {passed_sheets}")
print(f"  Failed: {failed_sheets}")
print(f"  Success rate: {(passed_sheets/total_sheets)*100:.1f}%")
