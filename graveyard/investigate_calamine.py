"""Investigate what determines calamine's data boundaries"""

from python_calamine import CalamineWorkbook
import openpyxl

file_path = "sample_data/sample.xlsx"

print("=" * 80)
print("INVESTIGATING SUMMARY SHEET")
print("=" * 80)

# Calamine
print("\n1. CALAMINE:")
wb_cal = CalamineWorkbook.from_path(file_path)
sheet_cal = wb_cal.get_sheet_by_name("Summary")
data = sheet_cal.to_python()

print(f"   Total rows returned: {len(data)}")
print(f"   Row 65 (index 64) length: {len(data[64])}")
print(f"   Row 66 exists in calamine data: {65 < len(data)}")

# Check what's in the last few rows
print(f"\n   Last 3 rows from calamine:")
for i in range(max(0, len(data)-3), len(data)):
    non_empty_count = sum(1 for cell in data[i] if cell is not None and cell != '')
    print(f"   Row {i+1} (index {i}): length={len(data[i])}, non_empty cells={non_empty_count}")

# Openpyxl
print("\n2. OPENPYXL:")
wb_xl = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
sheet_xl = wb_xl["Summary"]

print(f"   max_row property: {sheet_xl.max_row}")
print(f"   max_column property: {sheet_xl.max_column}")

# Check specific cells we know have data
print(f"\n3. SPECIFIC CELLS:")
print(f"   C66 value: {repr(sheet_xl['C66'].value)}")
print(f"   H66 value: {repr(sheet_xl['H66'].value)}")
print(f"   BW40 value: {repr(sheet_xl['BW40'].value)}")
print(f"   BW41 value: {repr(sheet_xl['BW41'].value)}")
print(f"   BW42 value: {repr(sheet_xl['BW42'].value)}")

# Check calamine's sheet properties
print(f"\n4. CALAMINE SHEET PROPERTIES:")
print(f"   sheet.height: {sheet_cal.height}")
print(f"   sheet.width: {sheet_cal.width}")
print(f"   sheet.total_height: {sheet_cal.total_height}")
print(f"   sheet.total_width: {sheet_cal.total_width}")
print(f"   sheet.start: {sheet_cal.start}")
print(f"   sheet.end: {sheet_cal.end}")

wb_xl.close()

print("\n" + "=" * 80)
print("CONCLUSION:")
print("=" * 80)
print("Calamine appears to determine boundaries based on Excel's internal")
print("'dimension' or 'used range' metadata, which may not include all cells")
print("with data, especially if there are gaps in the data.")
