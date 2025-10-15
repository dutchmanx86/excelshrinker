"""Investigate what Calamine returns for the Renewed sheet"""

from python_calamine import CalamineWorkbook
import openpyxl

file_path = "sample_data/sample.xlsx"
sheet_name = "Renewed"

# Expected: N7899 (column N = column 13, row 7899)
expected_col = 13  # 0-indexed (N is the 14th column)
expected_row = 7898  # 0-indexed

print("=" * 80)
print("INVESTIGATING RENEWED SHEET")
print("=" * 80)
print(f"Expected boundaries: Column N (index {expected_col}), Row 7899 (index {expected_row})")
print()

# CALAMINE
print("CALAMINE DATA:")
print("-" * 80)
workbook = CalamineWorkbook.from_path(file_path)
sheet = workbook.get_sheet_by_name(sheet_name)
data = sheet.to_python()

print(f"  Total rows returned by Calamine: {len(data)}")
print(f"  Max row length (columns): {max(len(row) for row in data) if data else 0}")
print()

# Check if the expected data exists in what Calamine returned
print("Checking if expected boundary data exists in Calamine's data:")
print(f"  Row {expected_row} (7899): ", end="")
if expected_row < len(data):
    print(f"EXISTS in Calamine data")
    print(f"    Row length: {len(data[expected_row])}")
    if expected_col < len(data[expected_row]):
        print(f"    Cell value at column {expected_col} (N): {data[expected_row][expected_col]}")
    else:
        print(f"    Column {expected_col} (N) NOT in row data (row only has {len(data[expected_row])} columns)")
else:
    print(f"MISSING - Calamine only returned {len(data)} rows")
print()

# Check the last few rows Calamine gave us
print("Last 5 rows from Calamine:")
for i in range(max(0, len(data) - 5), len(data)):
    row_data = data[i]
    non_empty = sum(1 for cell in row_data if cell is not None and str(cell).strip() != '')
    print(f"  Row {i + 1} (0-idx={i}): {len(row_data)} cols, {non_empty} non-empty cells")
print()

# OPENPYXL - Check the actual cells
print("OPENPYXL VERIFICATION:")
print("-" * 80)
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
sheet_xl = wb[sheet_name]

print(f"  openpyxl max_row: {sheet_xl.max_row}")
print(f"  openpyxl max_column: {sheet_xl.max_column}")
print()

# Check specific cells around the expected boundary
print(f"Checking cells around expected boundary (row 7899, column N):")
for row in [7897, 7898, 7899, 7900]:
    cell_val = sheet_xl.cell(row=row, column=14).value  # Column N is 14 (1-indexed)
    print(f"  Cell N{row} (row={row}, col=14): {repr(cell_val)}")

print()

# Scan for the last non-empty cell in column N
print("Scanning column N for last non-empty cell:")
last_non_empty_row = None
for row in range(1, 8000):
    cell_val = sheet_xl.cell(row=row, column=14).value
    if cell_val is not None and str(cell_val).strip() != '':
        last_non_empty_row = row

print(f"  Last non-empty cell in column N: Row {last_non_empty_row}")

wb.close()
