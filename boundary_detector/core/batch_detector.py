"""Optimized batch detection for multiple sheets in a single workbook"""

import time
from typing import Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from ..models.detection_config import DetectionConfig
from ..models.boundary_result import BoundaryResult, ChartInfo
from ..exceptions.errors import FileLoadError, DetectionError
from ..utils.excel_utils import is_cell_empty
from .chart_detector import ChartDetector


class BatchBoundaryDetector:
    """
    Optimized detector for analyzing multiple sheets in a single workbook.
    
    Opens the workbook once and reuses it for all sheets, dramatically
    reducing overhead when processing large workbooks.
    """
    
    def __init__(self, config: Optional[DetectionConfig] = None):
        """
        Initialize the batch detector.
        
        Args:
            config: Detection configuration. Uses defaults if not provided.
        """
        self.config = config or DetectionConfig()
    
    def detect_all_sheets(
        self,
        file_path: str,
        detect_charts: bool = True,
        parallel: bool = False,
        max_workers: int = 4
    ) -> Dict[str, BoundaryResult]:
        """
        Detect boundaries in all sheets of a workbook.
        
        This is much faster than calling detect() for each sheet separately,
        as it opens the workbook only once.
        
        Args:
            file_path: Path to the Excel file
            detect_charts: Whether to detect charts (default: True)
                          Set to False to skip chart detection and save ~30s
            parallel: Whether to process sheets in parallel (default: False)
            max_workers: Maximum parallel workers if parallel=True
        
        Returns:
            Dictionary mapping sheet names to BoundaryResult objects
        
        Raises:
            FileLoadError: If file cannot be loaded
            DetectionError: If detection process fails
        """
        start_time = time.time()
        
        # Validate file exists
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileLoadError(file_path, f"File not found: {file_path}")
        
        try:
            from python_calamine import CalamineWorkbook
        except ImportError:
            raise DetectionError(
                "python-calamine is not installed. "
                "Install it with: pip install python-calamine"
            )
        
        try:
            # Load calamine workbook (fast)
            print(f"Loading workbook with calamine: {file_path}")
            workbook = CalamineWorkbook.from_path(file_path)
            sheet_names = workbook.sheet_names
            print(f"Found {len(sheet_names)} sheets")
            print(f"Loading complete! ({(time.time() - start_time):.2f}s)")
            
            results = {}
            
            if parallel and len(sheet_names) > 1:
                # Parallel processing
                print(f"Processing sheets in parallel (max {max_workers} workers)...")
                results = self._process_sheets_parallel(
                    workbook,
                    file_path,
                    sheet_names,
                    detect_charts,
                    max_workers
                )
            else:
                # Sequential processing
                print("Processing sheets sequentially...")
                results = self._process_sheets_sequential(
                    workbook,
                    file_path,
                    sheet_names,
                    detect_charts
                )
            
            total_time = (time.time() - start_time) * 1000
            print(f"Total processing time: {total_time:.2f}ms ({total_time/1000:.2f}s)")
            print(f"Average per sheet: {total_time/len(sheet_names):.2f}ms")
            
            return results
            
        except Exception as e:
            raise DetectionError(f"Error during batch detection: {str(e)}") from e
    
    def _process_sheets_sequential(
        self,
        workbook,
        file_path: str,
        sheet_names: List[str],
        detect_charts: bool = True
    ) -> Dict[str, BoundaryResult]:
        """Process sheets one at a time, optionally detecting charts"""
        results = {}
        
        # Optionally load openpyxl workbook for chart detection
        openpyxl_wb = None
        if detect_charts:
            # NOTE: Must use normal mode (not read_only) to access charts
            # This is slow (~30s for large files) but only happens once
            print("Loading workbook for chart detection (this may take 20-30s for large files)...")
            chart_load_start = time.time()
            try:
                from openpyxl import load_workbook
                openpyxl_wb = load_workbook(file_path, data_only=True)
                print(f"Chart detection ready! ({(time.time() - chart_load_start):.2f}s)")
            except Exception as e:
                print(f"Warning: Could not load workbook for chart detection: {e}")
                openpyxl_wb = None
        else:
            print("Skipping chart detection (detect_charts=False)")
        
        try:
            for i, sheet_name in enumerate(sheet_names, 1):
                print(f"  [{i}/{len(sheet_names)}] {sheet_name}...", end=' ')
                try:
                    start = time.time()
                    
                    # Get sheet data from already-loaded calamine workbook
                    sheet = workbook.get_sheet_by_name(sheet_name)
                    
                    # Scan boundaries
                    boundaries, stats = self._scan_sheet(sheet)
                    
                    # Detect charts from already-loaded openpyxl workbook
                    charts = self._extract_charts_from_workbook(openpyxl_wb, sheet_name)
                    
                    # Check overlaps
                    max_col, max_row = boundaries
                    for chart in charts:
                        chart.overlaps_data = ChartDetector.check_overlap(
                            chart, max_col, max_row
                        )
                    
                    processing_time = (time.time() - start) * 1000
                    
                    results[sheet_name] = BoundaryResult(
                        max_column=max_col,
                        max_row=max_row,
                        total_columns=max_col + 1,
                        total_rows=max_row + 1,
                        charts=charts,
                        sheet_name=sheet_name,
                        processing_time_ms=processing_time,
                        cells_scanned=stats['cells_scanned'],
                        non_empty_cells=stats['non_empty_cells']
                    )
                    
                    print(f"OK {processing_time:.0f}ms")

                except Exception as e:
                    print(f"ERROR: {str(e)[:50]}")
                    continue
        finally:
            if openpyxl_wb:
                openpyxl_wb.close()
        
        return results
    
    def _process_sheets_parallel(
        self,
        workbook,
        file_path: str,
        sheet_names: List[str],
        detect_charts: bool,
        max_workers: int
    ) -> Dict[str, BoundaryResult]:
        """
        Process sheets in parallel.
        
        Note: Due to Rust borrow checker, we can't share the calamine workbook.
        Each thread needs to load its own calamine data, but we share the openpyxl workbook.
        """
        results = {}
        
        # Optionally load openpyxl workbook for chart detection
        openpyxl_wb = None
        if detect_charts:
            # NOTE: Must use normal mode (not read_only) to access charts
            # This is slow (~30s for large files) but only happens once
            print("Loading workbook for chart detection (this may take 20-30s for large files)...")
            chart_load_start = time.time()
            try:
                from openpyxl import load_workbook
                openpyxl_wb = load_workbook(file_path, data_only=True)
                print(f"Chart detection ready! ({(time.time() - chart_load_start):.2f}s)")
            except Exception as e:
                print(f"Warning: Could not load workbook for chart detection: {e}")
                openpyxl_wb = None
        else:
            print("Skipping chart detection (detect_charts=False)")
        
        try:
            def process_sheet(sheet_name: str) -> tuple:
                """Process a single sheet"""
                try:
                    start = time.time()
                    
                    # Each thread gets its own sheet data from main workbook
                    # (We pass the data, not the workbook object)
                    from python_calamine import CalamineWorkbook
                    wb = CalamineWorkbook.from_path(file_path)
                    sheet = wb.get_sheet_by_name(sheet_name)
                    
                    # Scan boundaries
                    boundaries, stats = self._scan_sheet(sheet)
                    
                    # Detect charts from shared openpyxl workbook
                    charts = self._extract_charts_from_workbook(openpyxl_wb, sheet_name)
                    
                    # Check overlaps
                    max_col, max_row = boundaries
                    for chart in charts:
                        chart.overlaps_data = ChartDetector.check_overlap(
                            chart, max_col, max_row
                        )
                    
                    processing_time = (time.time() - start) * 1000
                    
                    result = BoundaryResult(
                        max_column=max_col,
                        max_row=max_row,
                        total_columns=max_col + 1,
                        total_rows=max_row + 1,
                        charts=charts,
                        sheet_name=sheet_name,
                        processing_time_ms=processing_time,
                        cells_scanned=stats['cells_scanned'],
                        non_empty_cells=stats['non_empty_cells']
                    )
                    
                    return (sheet_name, result, None)
                    
                except Exception as e:
                    return (sheet_name, None, str(e))
            
            # Submit all tasks
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(process_sheet, name): name
                    for name in sheet_names
                }
                
                # Collect results as they complete
                completed = 0
                for future in as_completed(futures):
                    sheet_name, result, error = future.result()
                    completed += 1
                    
                    if error:
                        print(f"  [{completed}/{len(sheet_names)}] {sheet_name}: ERROR: {error[:50]}")
                    else:
                        results[sheet_name] = result
                        print(f"  [{completed}/{len(sheet_names)}] {sheet_name}: OK {result.processing_time_ms:.0f}ms")
        finally:
            if openpyxl_wb:
                openpyxl_wb.close()
        
        return results
    
    def _scan_sheet(self, sheet) -> tuple:
        """
        Scan a single sheet for boundaries.
        
        Args:
            sheet: Calamine sheet object (already loaded)
        
        Returns:
            Tuple of ((max_col, max_row), statistics_dict)
        """
        # Get all data at once (calamine loads efficiently)
        data = sheet.to_python()
        
        # Initialize tracking variables
        max_col_with_data = 0
        max_row_with_data = 0
        consecutive_empty_rows = 0
        cells_scanned = 0
        non_empty_cells = 0
        
        # Scan rows
        for current_row, row_data in enumerate(data):
            # Check if we can terminate early
            if (current_row >= self.config.min_rows and 
                consecutive_empty_rows >= self.config.empty_threshold):
                break
            
            consecutive_empty_cells = 0
            row_has_data = False
            
            # Determine how far to scan in this row
            max_col_to_scan = max(
                self.config.min_columns,
                max_col_with_data + self.config.empty_threshold
            )
            
            # Scan cells in this row
            for current_col in range(max_col_to_scan):
                cells_scanned += 1
                
                # Get cell value (or None if beyond row length)
                cell_value = row_data[current_col] if current_col < len(row_data) else None
                
                if is_cell_empty(cell_value):
                    consecutive_empty_cells += 1
                    
                    # Early exit if past minimum and hit threshold
                    if (current_col >= self.config.min_columns and 
                        consecutive_empty_cells >= self.config.empty_threshold):
                        break
                else:
                    consecutive_empty_cells = 0
                    max_col_with_data = max(max_col_with_data, current_col)
                    row_has_data = True
                    non_empty_cells += 1
            
            # Update row tracking
            if row_has_data:
                max_row_with_data = current_row
                consecutive_empty_rows = 0
            else:
                consecutive_empty_rows += 1
        
        # Compile statistics
        stats = {
            'cells_scanned': cells_scanned,
            'non_empty_cells': non_empty_cells
        }
        
        return (max_col_with_data, max_row_with_data), stats
    
    def _extract_charts_from_workbook(self, openpyxl_wb, sheet_name: str) -> List[ChartInfo]:
        """
        Extract charts from an already-loaded openpyxl workbook.
        
        Args:
            openpyxl_wb: Already-loaded openpyxl workbook object
            sheet_name: Sheet name to extract charts from
        
        Returns:
            List of ChartInfo objects
        """
        if openpyxl_wb is None:
            return []
        
        try:
            if sheet_name not in openpyxl_wb.sheetnames:
                return []
            
            worksheet = openpyxl_wb[sheet_name]
            charts = []
            
            try:
                if hasattr(worksheet, '_charts'):
                    for chart in worksheet._charts:
                        chart_info = ChartDetector._extract_chart_info(chart)
                        if chart_info:
                            charts.append(chart_info)
            except Exception:
                pass
            
            return charts
            
        except Exception:
            return []