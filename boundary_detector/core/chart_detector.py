"""Chart detection logic using openpyxl"""

from typing import List, Optional
from pathlib import Path

from ..models.boundary_result import ChartInfo
from ..exceptions.errors import FileLoadError, SheetNotFoundError


class ChartDetector:
    """
    Chart detector using openpyxl.
    
    This class handles chart detection separately from cell scanning,
    using openpyxl which is the only mature library with chart support.
    """
    
    @staticmethod
    def detect_charts(file_path: str, sheet_name: Optional[str] = None) -> List[ChartInfo]:
        """
        Detect all charts in a worksheet using openpyxl.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or None for first sheet
        
        Returns:
            List of ChartInfo objects
        
        Raises:
            FileLoadError: If file cannot be loaded
            SheetNotFoundError: If specified sheet is not found
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise FileLoadError(
                file_path,
                "openpyxl is not installed. Install it with: pip install openpyxl"
            )
        
        # Validate file exists
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileLoadError(file_path, f"File not found: {file_path}")
        
        try:
            # Load workbook in normal mode (read-only doesn't support charts)
            workbook = load_workbook(file_path, data_only=True)
            
            # Get the requested worksheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise SheetNotFoundError(sheet_name, workbook.sheetnames)
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            charts = []
            
            # Extract chart information
            # Note: In read_only mode, _charts might not be available
            # We need to handle this gracefully
            try:
                if hasattr(worksheet, '_charts'):
                    for chart in worksheet._charts:
                        chart_info = ChartDetector._extract_chart_info(chart)
                        if chart_info:
                            charts.append(chart_info)
            except Exception:
                # If chart detection fails, return empty list rather than failing
                # This makes the tool more robust
                pass
            
            workbook.close()
            return charts
            
        except SheetNotFoundError:
            raise
        except Exception as e:
            raise FileLoadError(file_path, f"Error reading charts with openpyxl: {str(e)}")
    
    @staticmethod
    def _extract_chart_info(chart) -> Optional[ChartInfo]:
        """
        Extract information from an openpyxl chart object.
        
        Args:
            chart: openpyxl chart object
        
        Returns:
            ChartInfo object or None if extraction fails
        """
        try:
            # Get chart anchor (position)
            anchor = chart.anchor
            
            # Extract position coordinates
            # anchor._from contains the top-left position
            # anchor.to contains the bottom-right position
            from_col = anchor._from.col if hasattr(anchor._from, 'col') else 0
            from_row = anchor._from.row if hasattr(anchor._from, 'row') else 0
            to_col = anchor.to.col if hasattr(anchor.to, 'col') else from_col
            to_row = anchor.to.row if hasattr(anchor.to, 'row') else from_row
            
            return ChartInfo(
                chart_type=chart.__class__.__name__,
                from_column=from_col,
                from_row=from_row,
                to_column=to_col,
                to_row=to_row,
                overlaps_data=False  # Will be set by caller
            )
        except Exception:
            # If we can't extract chart info, skip it
            return None
    
    @staticmethod
    def check_overlap(chart: ChartInfo, max_col: int, max_row: int) -> bool:
        """
        Check if a chart overlaps with the detected data boundary.
        
        Args:
            chart: ChartInfo object
            max_col: Maximum column with data (0-indexed)
            max_row: Maximum row with data (0-indexed)
        
        Returns:
            True if chart overlaps with data area, False otherwise
        """
        # Chart overlaps if its top-left corner is within the data bounds
        # or if it spans across the data area
        return (
            chart.from_column <= max_col and chart.from_row <= max_row
        ) or (
            chart.to_column >= 0 and chart.to_row >= 0 and
            chart.from_column <= max_col and chart.from_row <= max_row
        )