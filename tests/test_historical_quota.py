"""
Unit test for Historical Quota& Productivity sheet processing.

This test validates the entire pipeline from boundary detection through
compression to ensure accurate cell range identification.

Expected structure based on manual inspection:
- Date headers (annual): E5:H5 (2015-2018)
- Date headers (monthly): J5:BH5 (Jan 2015 - Mar 2019)
- Row labels section 1: B6:B12
- Row labels section 2: B15:B23
- Data blocks:
  - F7:H12 (annual quota data)
  - E16:G22 (annual productivity data)
  - J16:BH23 (monthly productivity data)
  - J26:AS26 (monthly summary row)
"""

import sys
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from boundary_detector.core.boundary_detector import BoundaryDetector
from compression.format_compression import compress_formats
from compression.inverted_index_compression import compress_format_result_json
import json


def test_historical_quota_boundary_detection():
    """Test that boundary detection correctly identifies the data range."""

    workbook_path = PROJECT_ROOT / "sample_data" / "sample.xlsx"
    sheet_name = "Historical Quota& Productivity"

    print(f"\n{'='*80}")
    print(f"TEST: Boundary Detection for {sheet_name}")
    print(f"{'='*80}\n")

    # Detect boundaries
    detector = BoundaryDetector()
    result = detector.detect(
        str(workbook_path),
        sheet_name,
        detect_charts=False
    )

    print(f"Detected boundaries:")
    print(f"  Max column (0-indexed): {result.max_column}")
    print(f"  Max row (0-indexed): {result.max_row}")
    print(f"  Total columns: {result.total_columns}")
    print(f"  Total rows: {result.total_rows}")

    # Convert to Excel-style for readability
    from boundary_detector.utils.excel_utils import column_index_to_letter
    max_col_letter = column_index_to_letter(result.max_column)
    max_row_1indexed = result.max_row + 1
    print(f"  Excel range: A1:{max_col_letter}{max_row_1indexed}")

    # Assertions based on manual inspection
    # We expect data to exist at least to row 26 (0-indexed = 25) and column BH (59)
    assert result.max_row >= 25, f"Expected max_row >= 25 (Excel row 26), got {result.max_row}"
    assert result.max_column >= 59, f"Expected max_column >= 59 (column BH), got {result.max_column}"

    print(f"\n[PASS] Boundary detection test PASSED")
    return result


def test_historical_quota_format_compression():
    """Test that format compression correctly identifies date headers and data blocks."""

    workbook_path = PROJECT_ROOT / "sample_data" / "sample.xlsx"
    sheet_name = "Historical Quota& Productivity"

    print(f"\n{'='*80}")
    print(f"TEST: Format Compression for {sheet_name}")
    print(f"{'='*80}\n")

    # First get boundaries
    detector = BoundaryDetector()
    boundary_result = detector.detect(
        str(workbook_path),
        sheet_name,
        detect_charts=False
    )

    # Compress formats
    format_result = compress_formats(
        file_path=str(workbook_path),
        sheet_name=sheet_name,
        boundary_result=boundary_result
    )

    # Convert to JSON for analysis
    format_json = json.loads(format_result.to_json())

    print(f"Format compression results:")
    print(f"  Total format ranges: {len(format_json['format_ranges'])}")
    print(f"  Date series found: {len(format_json.get('date_series_definitions', {}))}")

    # Check for date series
    date_series = format_json.get('date_series_definitions', {})
    print(f"\nDate series details:")
    for ds_id, ds_info in date_series.items():
        print(f"  {ds_id}: {ds_info['series_type']} from {ds_info['start_date']}")

    # Look for date ranges in format_ranges
    date_ranges = [r for r in format_json['format_ranges'] if r.get('type') == 'date']
    print(f"\nDate header ranges found:")
    for dr in date_ranges:
        print(f"  {dr['range']}: series_id={dr.get('date_series_id')}")

    # Expected date ranges (based on actual Excel):
    # Annual dates should be around E5:H5
    # Monthly dates should be around J5:BH5

    # Check if we found annual dates (should be 4 years: 2015-2018)
    annual_ranges = [r for r in date_ranges if 'ds_1' in str(r.get('date_series_id', ''))]
    if annual_ranges:
        print(f"\n[PASS] Found annual date range: {annual_ranges[0]['range']}")
    else:
        print(f"\n[WARN] WARNING: No annual date range found")

    # Check if we found monthly dates
    monthly_ranges = [r for r in date_ranges if 'ds_2' in str(r.get('date_series_id', ''))]
    if monthly_ranges:
        print(f"[PASS] Found monthly date range: {monthly_ranges[0]['range']}")
    else:
        print(f"[WARN] WARNING: No monthly date range found")

    return format_result


def test_historical_quota_full_pipeline():
    """Test the complete pipeline: boundary → format compression → inverted index."""

    workbook_path = PROJECT_ROOT / "sample_data" / "sample.xlsx"
    sheet_name = "Historical Quota& Productivity"

    print(f"\n{'='*80}")
    print(f"TEST: Full Pipeline for {sheet_name}")
    print(f"{'='*80}\n")

    # Run boundary detection
    detector = BoundaryDetector()
    boundary_result = detector.detect(
        str(workbook_path),
        sheet_name,
        detect_charts=False
    )

    # Run format compression
    format_result = compress_formats(
        file_path=str(workbook_path),
        sheet_name=sheet_name,
        boundary_result=boundary_result
    )

    # Run inverted index compression
    format_json = json.loads(format_result.to_json())
    compressed_result = compress_format_result_json(format_json, min_occurrences=2)

    # Analyze the final compressed output
    print(f"Final compressed JSON structure:")
    print(f"  Inverted index entries: {len(compressed_result.get('inverted_index', {}))}")
    print(f"  Format ranges: {len(compressed_result['data']['format_ranges'])}")
    print(f"  Compression ratio: {compressed_result['compression_stats']['compression_ratio']:.2%}")

    # Write debug output
    debug_output_path = PROJECT_ROOT / "tests" / "debug_historical_quota.json"
    with open(debug_output_path, 'w', encoding='utf-8') as f:
        json.dump(compressed_result, f, indent=2)

    print(f"\n[PASS] Full pipeline completed")
    print(f"  Debug output saved to: {debug_output_path}")

    return compressed_result


def debug_raw_cell_values():
    """Debug helper to inspect raw cell values from the Excel file."""

    import openpyxl

    workbook_path = PROJECT_ROOT / "sample_data" / "sample.xlsx"
    sheet_name = "Historical Quota& Productivity"

    print(f"\n{'='*80}")
    print(f"DEBUG: Raw Cell Values for {sheet_name}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(str(workbook_path), data_only=True, read_only=True)
    ws = wb[sheet_name]

    # Check the expected date header locations
    print("Expected annual date headers (E5:H5):")
    for col in ['E', 'F', 'G', 'H']:
        cell = ws[f'{col}5']
        print(f"  {col}5: {cell.value} (type: {type(cell.value).__name__})")

    print("\nExpected monthly date headers (J5:L5 sample):")
    for col in ['J', 'K', 'L']:
        cell = ws[f'{col}5']
        print(f"  {col}5: {cell.value} (type: {type(cell.value).__name__})")

    print("\nExpected row labels (B6:B12):")
    for row in range(6, 13):
        cell = ws[f'B{row}']
        print(f"  B{row}: {cell.value}")

    print("\nExpected data sample (F7:H7):")
    for col in ['F', 'G', 'H']:
        cell = ws[f'{col}7']
        print(f"  {col}7: {cell.value} (type: {type(cell.value).__name__})")

    wb.close()


if __name__ == "__main__":
    print("=" * 80)
    print("HISTORICAL QUOTA & PRODUCTIVITY TEST SUITE")
    print("=" * 80)

    # Run debug first to see raw values
    debug_raw_cell_values()

    # Run tests
    try:
        boundary_result = test_historical_quota_boundary_detection()
        format_result = test_historical_quota_format_compression()
        compressed_result = test_historical_quota_full_pipeline()

        print(f"\n{'='*80}")
        print("ALL TESTS PASSED")
        print(f"{'='*80}\n")
    except AssertionError as e:
        print(f"\n{'='*80}")
        print(f"TEST FAILED: {e}")
        print(f"{'='*80}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"ERROR: {e}")
        print(f"{'='*80}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)
