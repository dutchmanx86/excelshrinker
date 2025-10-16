"""
Excel Query Extractor
Queries Gemini for relevant cell ranges and extracts data from Excel files.
Outputs both raw JSON responses and extracted pandas DataFrames for auditing.
"""
import os
import json
import re
import argparse
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import google.generativeai as genai
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# --- Configuration ---
API_KEY = os.getenv("GEMINI_API_KEY")
MODEL_NAME = "gemini-2.5-pro"
BASE_DIR = Path(__file__).parent
DATASETS_DIR = BASE_DIR / "datasets"

BASE_PROMPT = """
You are an expert data-analysis lookup assistant. You will receive a markdown document that describes an Excel model: sheets, sections (e.g., "Income Statement"), and time-series ranges.

Retrieval Objective
Return cell ranges for the smallest valid block that actually exists in the document, prioritizing summary output sheets. When a user asks for a metric that belongs to a section (e.g., Income Statement)
and the markdown does not specify row-level line items, you must:
 - Retrieve the entire section block that contains the metric,
 - Clip that block to the requested date range (with buffer), and
 - Do not invent row labels or sub-ranges that are not explicitly documented.

Absolutely Do / Never Do
 - Do return block-level ranges when labels are unspecified.
 - Do include three ranges per answer:
   * Row Label Range
   * Column Label Range
   * Data Range: the interior data cells of the section.
 - Never infer or fabricate specific line items (e.g., "Gross Margin") unless the markdown explicitly lists them.
 - Never narrow to a single row if the markdown only names a section.

Time Series Rules (with Buffer)
 - Detect the section's full time axis (e.g., T2:FS2) and its frequency + start (e.g., "Monthly starting Jan 2015").
 - There may be multiple time series of data (e.g., Annual, Quarterly, Monthly). **Select the time granularity that best matches the user's query**
   * Use anchor points (if available) to accurately locate specific dates within long time series
 - Compute the column/row subset for the requested date window.
 - Add a buffer of two periods before and after.
 - Clip Row Label Range, Column Label Range, and Data Range to that subset.
 - If the requested window partially falls outside the available series, return the intersecting subset and note the truncation in a one-line remark.

Orientation & Section Boundary Detection
 - Orientation: Determine if the time axis is columns (dates across a header row) or rows (dates down a header column). Choose ranges accordingly.
 - Section boundaries: A section starts at its named header (e.g., "Income Statement") and ends at the next section header or a clearly defined boundary in the markdown. Use that rectangle as the section block.
 - If the markdown gives only a sheet name + section name (no cell coords), return the sheet + computed block based on the section's described span and time axis.

Output Format (max 3 sections)

Return at most the 3 most relevant sections. For each, output only the following JSON object:

{
  "sheet": "SheetName",
  "section": "SectionName",
  "row_label_range": "A6:A40",
  "column_label_range": "F3:K3",
  "data_range": "F6:K40",
  "full_time_range": "C3:N3",
  "notes": "Clipped to requested window + 2-period buffer. No row-level labels specified; returned full section block."
}

Selection Rules
 - Summary sheets first (e.g., Executive Summary, Consolidated Summary).
 - If multiple matching sections exist, pick the highest aggregation level unless the query specifies otherwise.
 - If the query names a specific line item and the markdown lists it, return its row sub-range; otherwise, fall back to the section block.

Sanity Checks Before Responding
 - If row or column labels are not explicitly provided, confirm notes explains that you returned the full section block without fabricated labels.
 - Verify ranges are contiguous rectangles and align with the orientation.

Ensure time clipping + buffer is reflected in all ranges.
"""


def find_excel_file(dataset_dir: Path) -> Optional[Path]:
    """
    Find the Excel file in the dataset's raw_data folder.

    Returns:
        Path to Excel file or None if not found
    """
    raw_data_dir = dataset_dir / "raw_data"
    if not raw_data_dir.exists():
        return None

    excel_files = list(raw_data_dir.glob("*.xlsx")) + list(raw_data_dir.glob("*.xls"))
    return excel_files[0] if excel_files else None


def find_markdown_file(dataset_dir: Path) -> Optional[Path]:
    """
    Find the MASTER_WORKBOOK_ANALYSIS.md file in the dataset's processing_outputs folder.

    Returns:
        Path to markdown file or None if not found
    """
    markdown_file = dataset_dir / "processing_outputs" / "MASTER_WORKBOOK_ANALYSIS.md"
    return markdown_file if markdown_file.exists() else None


def build_sheet_index(markdown_content: str) -> Dict[str, Dict[str, str]]:
    """
    Build a lightweight index of all sheets in the markdown file.

    Each sheet entry contains:
    - sheet_name: Name of the sheet
    - summary: The important parts of the sheet description
    - sections: List of section names and metrics contained if known
    - time_series: Time series ranges information if available

    Args:
        markdown_content: Full content of MASTER_WORKBOOK_ANALYSIS.md

    Returns:
        Dictionary mapping sheet names to their index entries
    """
    index = {}

    # Split by sheet headers (lines starting with ## Sheet:)
    sheet_sections = re.split(r'\n## Sheet: ', markdown_content)

    for section in sheet_sections[1:]:  # Skip first split (before any sheets)
        lines = section.split('\n')
        sheet_name = lines[0].strip()

        # Extract summary (first paragraph after sheet name)
        summary_lines = []
        sections_list = []
        time_series_info = ""

        in_summary = False
        for line in lines[1:]:
            # Stop at first section header
            if line.startswith('### '):
                section_title = line.replace('### ', '').strip()
                sections_list.append(section_title)
                in_summary = False
            elif line.startswith('**Purpose'):
                in_summary = True
            elif in_summary and line.strip():
                summary_lines.append(line.strip())
                if len(summary_lines) >= 3:  # Limit summary to 3 lines
                    in_summary = False
            elif 'Time Series' in line or 'anchor points' in line.lower():
                time_series_info += line.strip() + " "

        index[sheet_name] = {
            'sheet_name': sheet_name,
            'summary': ' '.join(summary_lines[:3]),  # First 3 lines
            'sections': sections_list,
            'time_series': time_series_info.strip()
        }

    return index


def query_index_for_relevant_sheets(
    index: Dict[str, Dict[str, str]],
    query: str,
    max_sheets: int = 5
) -> List[str]:
    """
    Query the sheet index to find the most relevant sheets using Gemini.

    Args:
        index: Sheet index from build_sheet_index()
        query: User's query
        max_sheets: Maximum number of sheets to return

    Returns:
        List of relevant sheet names
    """
    # Create a compact summary of all sheets
    index_summary = "# Sheet Index\n\n"
    for sheet_name, info in index.items():
        index_summary += f"## {sheet_name}\n"
        if info['summary']:
            index_summary += f"Summary: {info['summary']}\n"
        if info['sections']:
            index_summary += f"Sections: {', '.join(info['sections'][:5])}\n"  # First 5 sections
        if info['time_series']:
            index_summary += f"Time Series: {info['time_series'][:200]}\n"  # First 200 chars
        index_summary += "\n"

    # Query Gemini to identify relevant sheets
    genai.configure(api_key=API_KEY)
    model = genai.GenerativeModel(
        MODEL_NAME,
        generation_config={
            "temperature": 0.0,
            "top_p": 1.0,
            "top_k": 1
        }
    )

    prompt = f"""You are analyzing an Excel workbook index to find the most relevant sheets for a user query.

Here is the index of all sheets:

{index_summary}

User query: "{query}"

Return ONLY a JSON array of the {max_sheets} most relevant sheet names, in order of relevance.
Example: ["Sheet1", "Sheet2", "Sheet3"]

Sheet names:"""

    response = model.generate_content(prompt)

    # Parse the JSON response
    try:
        # Try to extract JSON array from response
        text = response.text.strip()
        # Remove markdown code blocks if present
        text = re.sub(r'```json\s*', '', text)
        text = re.sub(r'```\s*', '', text)
        sheet_names = json.loads(text)
        return sheet_names[:max_sheets]
    except:
        # Fallback: return all sheet names if parsing fails
        return list(index.keys())[:max_sheets]


def parse_json_responses(response_text: str) -> List[Dict]:
    """
    Parse JSON objects from Gemini response text.
    Extracts up to 3 JSON objects containing cell range information.

    Returns:
        List of parsed JSON dictionaries
    """
    json_objects = []

    # Try to find JSON code blocks first
    json_blocks = re.findall(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)

    if json_blocks:
        for block in json_blocks[:3]:  # Max 3
            try:
                json_obj = json.loads(block)
                json_objects.append(json_obj)
            except json.JSONDecodeError:
                continue
    else:
        # Try to find raw JSON objects
        json_matches = re.findall(r'\{[^{}]*"sheet"[^{}]*"section"[^{}]*\}', response_text, re.DOTALL)
        for match in json_matches[:3]:  # Max 3
            try:
                json_obj = json.loads(match)
                json_objects.append(json_obj)
            except json.JSONDecodeError:
                continue

    return json_objects


def parse_range(range_str: str) -> Tuple[str, int, str, int]:
    """
    Parse an Excel range string like "A6:B10" into components.

    Returns:
        Tuple of (start_col, start_row, end_col, end_row)
    """
    match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str)
    if not match:
        raise ValueError(f"Invalid range format: {range_str}")

    start_col, start_row, end_col, end_row = match.groups()
    return start_col, int(start_row), end_col, int(end_row)


def col_to_index(col_letter: str) -> int:
    """Convert Excel column letter to 0-based index (A=0, B=1, etc.)"""
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def extract_range_from_sheet(sheet, range_str: str) -> List[List]:
    """
    Extract cell values from a sheet given a range string.

    Args:
        sheet: openpyxl worksheet object
        range_str: Excel range like "A6:B10"

    Returns:
        2D list of cell values
    """
    start_col, start_row, end_col, end_row = parse_range(range_str)

    # Convert to indices
    start_col_idx = col_to_index(start_col)
    end_col_idx = col_to_index(end_col)

    # Extract values
    data = []
    for row_idx in range(start_row, end_row + 1):
        row_data = []
        for col_idx in range(start_col_idx, end_col_idx + 1):
            cell = sheet.cell(row=row_idx, column=col_idx + 1)  # openpyxl is 1-indexed
            row_data.append(cell.value)
        data.append(row_data)

    return data


def create_labeled_dataframe(
    row_labels: List[List],
    col_labels: List[List],
    data: List[List]
) -> pd.DataFrame:
    """
    Create a pandas DataFrame with row and column labels.

    Args:
        row_labels: 2D list from row_label_range (typically single column)
        col_labels: 2D list from column_label_range (typically single row)
        data: 2D list from data_range

    Returns:
        Labeled pandas DataFrame
    """
    # Flatten row labels (assuming single column)
    if row_labels and len(row_labels) > 0:
        row_index = [row[0] if row else None for row in row_labels]
    else:
        row_index = None

    # Flatten column labels (assuming single row)
    if col_labels and len(col_labels) > 0:
        col_index = col_labels[0]
    else:
        col_index = None

    # TEMPORARY FIX: Pad labels or data to match dimensions
    # This handles cases where Gemini returns misaligned row/column label ranges
    # TODO: Update prompt to ensure label ranges align exactly with data ranges
    if row_index is not None and len(data) > 0:
        if len(row_index) < len(data):
            # Pad row labels with None
            row_index.extend([None] * (len(data) - len(row_index)))
        elif len(row_index) > len(data):
            # Pad data rows with None values
            num_cols = len(data[0]) if data else 0
            for _ in range(len(row_index) - len(data)):
                data.append([None] * num_cols)

    if col_index is not None and len(data) > 0 and len(data[0]) > 0:
        if len(col_index) < len(data[0]):
            # Pad column labels with None
            col_index.extend([None] * (len(data[0]) - len(col_index)))
        elif len(col_index) > len(data[0]):
            # Pad data columns with None values
            for row in data:
                row.extend([None] * (len(col_index) - len(row)))

    # Create DataFrame
    df = pd.DataFrame(data, index=row_index, columns=col_index)

    return df


def extract_data_for_range(
    workbook_path: Path,
    range_json: Dict
) -> Tuple[pd.DataFrame, Dict]:
    """
    Extract data from Excel file for a given range JSON.

    Args:
        workbook_path: Path to Excel file
        range_json: JSON object with sheet and range information

    Returns:
        Tuple of (DataFrame, extraction_info dict)
    """
    wb = load_workbook(str(workbook_path), data_only=True)
    sheet_name = range_json['sheet']

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    sheet = wb[sheet_name]

    # Extract ranges
    row_labels = extract_range_from_sheet(sheet, range_json['row_label_range'])
    col_labels = extract_range_from_sheet(sheet, range_json['column_label_range'])
    data = extract_range_from_sheet(sheet, range_json['data_range'])

    # Create DataFrame
    df = create_labeled_dataframe(row_labels, col_labels, data)

    # Extraction info for auditing
    info = {
        'sheet': sheet_name,
        'section': range_json['section'],
        'row_label_range': range_json['row_label_range'],
        'column_label_range': range_json['column_label_range'],
        'data_range': range_json['data_range'],
        'notes': range_json.get('notes', ''),
        'dataframe_shape': df.shape
    }

    wb.close()

    return df, info


def extract_relevant_sheets(markdown_content: str, sheet_names: List[str]) -> str:
    """
    Extract only the relevant sheet sections from the full markdown content.

    Args:
        markdown_content: Full MASTER_WORKBOOK_ANALYSIS.md content
        sheet_names: List of sheet names to extract

    Returns:
        Filtered markdown containing only the specified sheets
    """
    # Split by sheet headers
    sheet_sections = re.split(r'\n## Sheet: ', markdown_content)

    # Keep the header (before first sheet)
    filtered_content = sheet_sections[0]

    # Add only the requested sheets
    for section in sheet_sections[1:]:
        lines = section.split('\n')
        sheet_name = lines[0].strip()

        if sheet_name in sheet_names:
            filtered_content += f"\n## Sheet: {section}"

    return filtered_content


def query_gemini_and_extract(
    query: str,
    markdown_content: str,
    workbook_path: Path,
    use_index: bool = True
) -> Tuple[str, List[Dict], List[pd.DataFrame]]:
    """
    Query Gemini for relevant ranges and extract data from Excel.
    Uses a two-pass approach: first queries an index to find relevant sheets,
    then sends only those sheets to Gemini for detailed analysis.

    Args:
        query: User's question
        markdown_content: Content of MASTER_WORKBOOK_ANALYSIS.md
        workbook_path: Path to Excel file
        use_index: If True, use two-pass approach with index. If False, send full markdown.

    Returns:
        Tuple of (raw_response, json_ranges, dataframes)
    """
    # PASS 1: Query index to find relevant sheets (if enabled)
    if use_index:
        print("  Building sheet index...")
        index = build_sheet_index(markdown_content)
        print(f"  Found {len(index)} sheets in index")

        print("  Querying index for relevant sheets...")
        relevant_sheets = query_index_for_relevant_sheets(index, query, max_sheets=5)
        print(f"  Selected {len(relevant_sheets)} relevant sheets: {', '.join(relevant_sheets)}")

        # Extract only relevant sheets from markdown
        filtered_markdown = extract_relevant_sheets(markdown_content, relevant_sheets)
        print(f"  Reduced markdown from {len(markdown_content)} to {len(filtered_markdown)} chars")
    else:
        filtered_markdown = markdown_content

    # PASS 2: Query Gemini with filtered markdown for detailed range extraction
    print("  Querying Gemini for cell ranges...")
    genai.configure(api_key=API_KEY)
    model = genai.GenerativeModel(
        MODEL_NAME,
        generation_config={
            "temperature": 0.0,  # Make responses deterministic
            "top_p": 1.0,
            "top_k": 1
        }
    )
    chat = model.start_chat(history=[])

    # Send initial context
    initial_prompt = f"{BASE_PROMPT}\n\nHere is the content of the markdown file:\n\n{filtered_markdown}"
    chat.send_message(initial_prompt)

    # Send user query
    user_prompt = f"The user's query is: '{query}'. Find the 3 most relevant ranges as per the base prompt, please."
    response = chat.send_message(user_prompt)

    raw_response = response.text

    # Parse JSON responses
    json_ranges = parse_json_responses(raw_response)

    # Extract data for each range
    print("  Extracting data from Excel...")
    dataframes = []
    for range_json in json_ranges:
        try:
            df, info = extract_data_for_range(workbook_path, range_json)
            dataframes.append({
                'dataframe': df,
                'info': info
            })
        except Exception as e:
            print(f"  Error extracting data for range: {e}")
            dataframes.append({
                'dataframe': None,
                'info': {'error': str(e)}
            })

    return raw_response, json_ranges, dataframes


def print_audit_output(
    query: str,
    raw_response: str,
    json_ranges: List[Dict],
    dataframes: List[Dict]
):
    """
    Print audit output showing raw JSON and extracted DataFrames.
    """
    print("\n" + "="*80)
    print("EXCEL QUERY EXTRACTOR - AUDIT OUTPUT")
    print("="*80)

    print(f"\nQuery: {query}")

    print("\n" + "-"*80)
    print("RAW GEMINI RESPONSE:")
    print("-"*80)
    print(raw_response)

    print("\n" + "-"*80)
    print(f"PARSED JSON RANGES ({len(json_ranges)} found):")
    print("-"*80)
    for i, json_range in enumerate(json_ranges, 1):
        print(f"\n[Range {i}]")
        print(json.dumps(json_range, indent=2))

    print("\n" + "-"*80)
    print(f"EXTRACTED DATAFRAMES ({len(dataframes)} total):")
    print("-"*80)

    for i, df_info in enumerate(dataframes, 1):
        print(f"\n[DataFrame {i}]")

        if 'error' in df_info['info']:
            print(f"ERROR: {df_info['info']['error']}")
            continue

        info = df_info['info']
        print(f"Sheet: {info['sheet']}")
        print(f"Section: {info['section']}")
        print(f"Row Labels: {info['row_label_range']}")
        print(f"Column Labels: {info['column_label_range']}")
        print(f"Data Range: {info['data_range']}")
        print(f"Shape: {info['dataframe_shape']}")
        print(f"Notes: {info['notes']}")
        print("\nData Preview:")
        print(df_info['dataframe'])

    print("\n" + "="*80)


def main():
    """
    Main CLI function for Excel Query Extractor.
    """
    parser = argparse.ArgumentParser(
        description="Query Gemini for relevant Excel ranges and extract data",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "dataset",
        help="Name of the dataset folder (e.g., 'alphasense')"
    )
    parser.add_argument(
        "query",
        nargs='?',
        help="Query to ask (optional, will prompt if not provided)"
    )

    args = parser.parse_args()

    # Verify dataset exists
    dataset_dir = DATASETS_DIR / args.dataset
    if not dataset_dir.exists():
        print(f"ERROR: Dataset folder not found: {dataset_dir}")
        print("\nAvailable datasets:")
        if DATASETS_DIR.exists():
            for item in sorted(DATASETS_DIR.iterdir()):
                if item.is_dir():
                    print(f"  - {item.name}")
        return

    # Find Excel file
    excel_file = find_excel_file(dataset_dir)
    if not excel_file:
        print(f"ERROR: No Excel file found in {dataset_dir / 'raw_data'}")
        return

    # Find markdown file
    markdown_file = find_markdown_file(dataset_dir)
    if not markdown_file:
        print(f"ERROR: No MASTER_WORKBOOK_ANALYSIS.md found in {dataset_dir / 'processing_outputs'}")
        return

    print(f"Dataset: {args.dataset}")
    print(f"Excel file: {excel_file.name}")
    print(f"Markdown file: {markdown_file.name}")

    # Load markdown content
    with open(markdown_file, 'r', encoding='utf-8') as f:
        markdown_content = f.read()

    # Get query
    if args.query:
        query = args.query
    else:
        query = input("\nEnter your query: ")

    if not query.strip():
        print("No query provided. Exiting.")
        return

    # Query Gemini and extract data
    print("\nQuerying Gemini and extracting data...")
    raw_response, json_ranges, dataframes = query_gemini_and_extract(
        query, markdown_content, excel_file
    )

    # Print audit output
    print_audit_output(query, raw_response, json_ranges, dataframes)


if __name__ == "__main__":
    main()
