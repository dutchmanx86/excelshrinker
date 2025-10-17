"""Format detection engine for Excel spreadsheets"""

import sys
import time
from typing import Optional, Tuple, List
from pathlib import Path
from datetime import datetime, date

from boundary_detector.models.detection_config import DetectionConfig
from boundary_detector.models.boundary_result import BoundaryResult
from boundary_detector.exceptions.errors import FileLoadError, SheetNotFoundError, DetectionError
from boundary_detector.utils.excel_utils import is_cell_empty, column_index_to_letter
from boundary_detector.core.boundary_detector import BoundaryDetector
from compression.format_detector.format_result import FormatResult
from compression.format_detector.format_range import FormatRange
from compression.format_detector.format_mappings import classify_format, get_format_scale
from compression.format_detector.time_period_parser import TimePeriodParser
from compression.format_detector.time_range_detector import TimeRangeDetector
from compression.format_detector.date_series_analyzer import DateSeriesAnalyzer


class FormatDetector:
    """
    High-performance Excel number format detector.

    Uses a hybrid approach:
    1. Calamine for fast cell value scanning
    2. Openpyxl for accurate number format code extraction
    """

    def __init__(self, config: Optional[DetectionConfig] = None):
        """
        Initialize the format detector.

        Args:
            config: Detection configuration. Uses defaults if not provided.
        """
        self.config = config or DetectionConfig()
        self.time_parser = TimePeriodParser()
        self.time_range_detector = TimeRangeDetector()

    @staticmethod
    def _safe_extract_color(color_obj) -> Optional[str]:
        """
        Safely extract color information from an openpyxl Color object.

        Args:
            color_obj: openpyxl.styles.colors.Color object or None

        Returns:
            String representation of the color, or None if no color
        """
        if not color_obj:
            return None

        # Check the color type to determine which attribute to access
        color_type = getattr(color_obj, 'type', None)

        if color_type == 'rgb':
            # RGB color - access rgb attribute
            rgb = getattr(color_obj, 'rgb', None)
            if rgb and rgb not in ('00000000', 'FFFFFFFF'):  # Skip default/white
                return f"rgb:{rgb}"

        elif color_type == 'theme':
            # Theme color - use theme index
            theme = getattr(color_obj, 'theme', None)
            tint = getattr(color_obj, 'tint', 0.0)
            if theme is not None and theme != 1:  # Skip default theme
                if tint != 0.0:
                    return f"theme:{theme}+{tint:.2f}"
                else:
                    return f"theme:{theme}"

        elif color_type == 'indexed':
            # Indexed color
            indexed = getattr(color_obj, 'indexed', None)
            if indexed is not None and indexed != 64:  # 64 is auto/default
                return f"indexed:{indexed}"

        return None

    @staticmethod
    def _safe_extract_fill_color(fill_obj) -> Optional[str]:
        """
        Safely extract background color from a fill object.

        Args:
            fill_obj: openpyxl fill object

        Returns:
            String representation of the background color, or None
        """
        if not fill_obj:
            return None

        pattern_type = getattr(fill_obj, 'patternType', None)
        if not pattern_type or pattern_type == 'none':
            return None

        # Try to get start_color
        start_color = getattr(fill_obj, 'start_color', None)
        if start_color:
            return FormatDetector._safe_extract_color(start_color)

        return None
    
    def detect(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        boundary_result: Optional[BoundaryResult] = None,
        verbose: bool = False
    ) -> FormatResult:
        """
        Detect number formats in an Excel worksheet.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to analyze. If None, uses first sheet.
            boundary_result: Optional pre-computed boundary result for efficiency
            verbose: If True, prints progress messages to stderr.
        
        Returns:
            FormatResult with detected format ranges
        
        Raises:
            FileLoadError: If file cannot be loaded
            SheetNotFoundError: If specified sheet is not found
            DetectionError: If detection process fails
        """
        start_time = time.time()
        
        # Validate file exists
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            raise FileLoadError(file_path, f"File not found: {file_path}")
        
        try:
            # Determine sheet name
            if sheet_name is None:
                sheet_name = self._get_first_sheet_name(file_path)
            
            # Get boundaries (either from parameter or detect using proper boundary detector)
            if boundary_result:
                max_col = boundary_result.max_column
                max_row = boundary_result.max_row
            else:
                # Use the proper BoundaryDetector which handles sparse data correctly
                boundary_detector = BoundaryDetector(self.config)
                boundary_result = boundary_detector.detect(file_path, sheet_name)
                max_col = boundary_result.max_column
                max_row = boundary_result.max_row
            
            # Scan formats using hybrid approach
            format_ranges, cells_analyzed, sheet_data, start_row, start_col, format_cache = self._scan_formats_hybrid(
                file_path, sheet_name, max_col, max_row, verbose=verbose
            )

            # Detect and consolidate vertical (columnar) time series using TimeRangeDetector
            columnar_ranges, consolidated_cells = self._detect_columnar_time_series(
                sheet_data, max_col, max_row, start_row, start_col, format_cache
            )

            # Detect and consolidate horizontal time series, excluding cells from columnar ranges
            time_series_ranges, time_series_cells = self._detect_and_consolidate_time_series(
                sheet_data, format_ranges, max_col, max_row, start_row, start_col, format_cache, exclude_cells=consolidated_cells
            )
            
            # Combine all non-date, non-time-series ranges with the new consolidated ranges
            final_ranges = []
            
            # Add all non-date ranges first
            for fmt_range in format_ranges:
                if fmt_range.type != 'date':
                    final_ranges.append(fmt_range)

            # Add the consolidated ranges
            final_ranges.extend(columnar_ranges)
            final_ranges.extend(time_series_ranges)

            # Remove any ranges that overlap with consolidated time series cells
            format_ranges = []
            all_consolidated_cells = consolidated_cells.union(time_series_cells)
            for fmt_range in final_ranges:
                # For single cells, check if they're in consolidated set
                if ':' not in fmt_range.range:
                    if fmt_range.range not in all_consolidated_cells:
                        format_ranges.append(fmt_range)
                else:
                    # For ranges, check if ANY cell overlaps with time series
                    # Only skip if it's a non-date range that overlaps
                    if fmt_range.type == 'date':
                        # Always keep date ranges (they ARE the time series)
                        format_ranges.append(fmt_range)
                    else:
                        # For non-date ranges, check if they overlap with time series
                        range_cells = self._get_cells_in_range(fmt_range.range)
                        if not any(cell in all_consolidated_cells for cell in range_cells):
                            # No overlap - keep the range
                            format_ranges.append(fmt_range)
                        # else: skip this range as it overlaps with a time series
            
            # Sort final ranges by row and then column position
            format_ranges.sort(key=lambda r: (self._get_row_from_range(r.range), self._get_col_from_range(r.range)))

            # Merge consecutive rows with identical column spans into rectangular regions
            format_ranges = self._merge_rectangular_regions(format_ranges)

            # Calculate processing time
            processing_time = (time.time() - start_time) * 1000  # Convert to ms
            
            # Create and return result
            result = FormatResult(
                sheet_name=sheet_name,
                format_ranges=format_ranges,
                processing_time_ms=processing_time,
                cells_analyzed=cells_analyzed
            )
            
            return result
            
        except FileLoadError:
            raise
        except SheetNotFoundError:
            raise
        except Exception as e:
            raise DetectionError(f"Error during format detection: {str(e)}") from e

    def _scan_formats_hybrid(
        self,
        file_path: str,
        sheet_name: str,
        max_col: int,
        max_row: int,
        verbose: bool = False
    ) -> Tuple[List[FormatRange], int, List[List], int, int]:
        """
        Scan formats using hybrid approach:
        1. Calamine for cell values (fast)
        2. Openpyxl for format codes (accurate, using iter_rows for speed)
        3. Group adjacent cells with same format

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            max_col: Maximum column to scan
            max_row: Maximum row to scan

        Returns:
            Tuple of (list of FormatRange objects, cells_analyzed count, sheet data, start_row, start_col, format_cache)
        """
        try:
            from python_calamine import CalamineWorkbook
            import openpyxl
        except ImportError as e:
            raise DetectionError(
                f"Required library not installed: {str(e)}. "
                "Install with: pip install python-calamine openpyxl"
            )

        # Load with calamine for calculated values (fast - handles formulas)
        if verbose:
            print("Loading workbook with calamine...", file=sys.stderr)
        workbook_cal = CalamineWorkbook.from_path(file_path)
        sheet_cal = workbook_cal.get_sheet_by_name(sheet_name)
        data = sheet_cal.to_python()

        # Get the offset from calamine - this tells us which Excel row/col data[0][0] corresponds to
        start_row, start_col = sheet_cal.start  # 0-indexed: (4, 1) means Excel row 5, column B
        if verbose:
            print(f"Calamine data starts at Excel row {start_row + 1}, column {start_col + 1}", file=sys.stderr)

        # Load with openpyxl for formats only - OPTIMIZED: use iter_rows
        if verbose:
            print("Loading formats with openpyxl (using iter_rows for speed)...", file=sys.stderr)
        workbook_xl = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        sheet_xl = workbook_xl[sheet_name]

        # PRE-LOAD all formats in one pass using iter_rows (MUCH faster than cell-by-cell)
        # Build cache for the ACTUAL range where calamine has data
        if verbose:
            print(f"Pre-loading format cache for {max_row + 1} rows x {max_col + 1} columns...", file=sys.stderr)
            print(f"  Fetching formats from Excel rows {start_row + 1} to {start_row + max_row + 1}", file=sys.stderr)
        format_cache = {}  # (row_idx, col_idx) -> (number_format, format_id, bold, italic, underline, font_color, bg_color)

        # Use iter_rows starting from where calamine's data actually begins
        for row_idx, row_cells in enumerate(sheet_xl.iter_rows(
            min_row=start_row + 1,  # Excel rows are 1-indexed
            max_row=start_row + max_row + 1,
            min_col=start_col + 1,
            max_col=start_col + max_col + 1
        )):
            for col_idx, cell_xl in enumerate(row_cells):
                number_format = cell_xl.number_format or 'General'
                format_id = cell_xl._style if hasattr(cell_xl, '_style') else 0

                # Extract styling attributes
                bold = bool(cell_xl.font and cell_xl.font.bold)
                italic = bool(cell_xl.font and cell_xl.font.italic)
                underline = bool(cell_xl.font and cell_xl.font.underline)

                # Extract font color (safe extraction)
                font_color = None
                if cell_xl.font and cell_xl.font.color:
                    font_color = self._safe_extract_color(cell_xl.font.color)

                # Extract background color (safe extraction)
                bg_color = self._safe_extract_fill_color(cell_xl.fill)

                # Cache key matches calamine data indices (0-indexed from start of data)
                format_cache[(row_idx, col_idx)] = (number_format, format_id, bold, italic, underline, font_color, bg_color)

        if verbose:
            print(f"Format cache loaded: {len(format_cache)} cells", file=sys.stderr)

        format_ranges = []
        cells_analyzed = 0

        # Progress tracking
        last_progress = 0

        if verbose:
            print("Scanning cells and grouping formats...", file=sys.stderr)
        # Scan row by row (top to bottom)
        for row_idx in range(max_row + 1):
            # Progress indicator every 10%
            if verbose and row_idx > 0 and (max_row + 1) > 100 and row_idx % ((max_row + 1) // 10) == 0:
                progress = (row_idx / (max_row + 1)) * 100
                if progress - last_progress >= 10:
                    print(f"Progress: {progress:.0f}% (row {row_idx}/{max_row + 1})", file=sys.stderr)
                    last_progress = progress

            current_range = None

            # Get row data from calamine (calculated values)
            row_data = data[row_idx] if row_idx < len(data) else []

            # Only scan up to max_col (not the full 16384 columns that calamine returns)
            scan_limit = min(max_col + 1, len(row_data))

            for col_idx in range(scan_limit):
                # Get cell value from calamine (handles formulas automatically)
                cell_value = row_data[col_idx] if col_idx < len(row_data) else None

                # Skip empty cells (don't even count them or access format cache)
                if is_cell_empty(cell_value):
                    # Save any current range
                    if current_range:
                        format_ranges.append(current_range)
                        current_range = None
                    continue

                # Only count and lookup format for non-empty cells
                cells_analyzed += 1

                # Get format from cache (instant lookup)
                number_format, format_id, bold, italic, underline, font_color, bg_color = format_cache.get(
                    (row_idx, col_idx), ('General', 0, False, False, False, None, None)
                )

                # Check value type and handle accordingly
                is_string_value = isinstance(cell_value, str)
                is_datetime_value = isinstance(cell_value, (datetime, date))

                if is_datetime_value:
                    # Cell contains a datetime object (from formula or direct date)
                    # Convert to ISO date string
                    if isinstance(cell_value, datetime):
                        parsed_date = cell_value.date().isoformat()
                    else:
                        parsed_date = cell_value.isoformat()

                    # Create single-cell range for datetime
                    # Convert calamine indices to Excel coordinates
                    excel_row = start_row + row_idx + 1  # Convert to 1-indexed Excel row
                    excel_col_letter = column_index_to_letter(start_col + col_idx)
                    cell_ref = f"{excel_col_letter}{excel_row}"
                    format_range = FormatRange(
                        range=cell_ref,
                        type='date',
                        parsed_date=parsed_date,
                        format_code=number_format
                    )
                    format_ranges.append(format_range)

                    # Dates don't continue ranges
                    if current_range:
                        format_ranges.append(current_range)
                        current_range = None

                elif is_string_value:
                    # Check if it's a time period string
                    if self.time_parser.is_time_period(cell_value):
                        parsed_date = self.time_parser.parse_to_date(cell_value)

                        # Create single-cell range for parsed time period
                        # Convert calamine indices to Excel coordinates
                        excel_row = start_row + row_idx + 1
                        excel_col_letter = column_index_to_letter(start_col + col_idx)
                        cell_ref = f"{excel_col_letter}{excel_row}"
                        format_range = FormatRange(
                            range=cell_ref,
                            type='date',
                            original_value=cell_value,
                            parsed_date=parsed_date,
                            format_code=number_format
                        )
                        format_ranges.append(format_range)
                    else:
                        # Regular string - create single-cell range
                        # Convert calamine indices to Excel coordinates
                        excel_row = start_row + row_idx + 1
                        excel_col_letter = column_index_to_letter(start_col + col_idx)
                        cell_ref = f"{excel_col_letter}{excel_row}"
                        format_range = FormatRange(
                            range=cell_ref,
                            type='string',
                            value=cell_value,
                            format_code=number_format
                        )
                        format_ranges.append(format_range)

                    # Strings don't continue ranges
                    if current_range:
                        format_ranges.append(current_range)
                        current_range = None

                else:
                    # Non-string value - classify format
                    format_type = classify_format(number_format, format_id)

                    # Group adjacent cells with same format
                    # Convert calamine indices to Excel coordinates
                    excel_row = start_row + row_idx + 1
                    excel_col_letter = column_index_to_letter(start_col + col_idx)
                    cell_ref = f"{excel_col_letter}{excel_row}"

                    if current_range is None:
                        # Start new range
                        current_range = self._create_range(
                            cell_ref, cell_ref, format_type, number_format, format_id,
                            bold=bold or None, italic=italic or None,
                            underline=underline or None, font_color=font_color, bg_color=bg_color
                        )
                    elif (current_range.type == format_type and
                          current_range.format_code == number_format and
                          current_range.bold == (bold or None) and
                          current_range.italic == (italic or None) and
                          current_range.underline == (underline or None) and
                          current_range.font_color == font_color and
                          current_range.bg_color == bg_color):
                        # Extend current range (format AND styling match)
                        current_range = self._create_range(
                            current_range.start_cell, cell_ref, format_type,
                            number_format, format_id,
                            bold=bold or None, italic=italic or None,
                            underline=underline or None, font_color=font_color, bg_color=bg_color
                        )
                    else:
                        # Different format or styling - save current and start new
                        format_ranges.append(current_range)
                        current_range = self._create_range(
                            cell_ref, cell_ref, format_type, number_format, format_id,
                            bold=bold or None, italic=italic or None,
                            underline=underline or None, font_color=font_color, bg_color=bg_color
                        )

            # End of row - save any open range
            if current_range:
                format_ranges.append(current_range)
                current_range = None

        workbook_xl.close()

        return format_ranges, cells_analyzed, data, start_row, start_col, format_cache

    def _detect_and_consolidate_time_series(
        self,
        data: List[List],
        format_ranges: List[FormatRange],
        max_col: int,
        max_row: int,
        start_row: int,
        start_col: int,
        format_cache: dict,
        exclude_cells: set = None
    ) -> Tuple[List[FormatRange], set]:
        """
        Detect horizontal time series patterns, excluding specified cells.

        Args:
            data: Sheet data from calamine.
            format_ranges: List of format ranges.
            max_col: Maximum column index.
            max_row: Maximum row index.
            start_row: Starting row offset from calamine (0-indexed).
            start_col: Starting column offset from calamine (0-indexed).
            format_cache: Dictionary mapping (row_idx, col_idx) to (number_format, format_id).
            exclude_cells: A set of cell references (e.g., "A1") to exclude.

        Returns:
            A tuple of (new time series ranges, set of cells in these series).
        """
        if exclude_cells is None:
            exclude_cells = set()

        time_series_cells = set()
        new_time_series_ranges = []

        for row_idx in range(max_row + 1):
            row_data = data[row_idx] if row_idx < len(data) else []

            # Extract format codes for this row
            row_formats = []
            for col_idx in range(len(row_data)):
                if (row_idx, col_idx) in format_cache:
                    number_format, *_ = format_cache[(row_idx, col_idx)]  # Unpack only number_format
                    row_formats.append(number_format)
                else:
                    row_formats.append(None)

            series_list = self.time_range_detector.detect_row_series(
                row_data, row_number=row_idx, start_col=0, format_codes=row_formats
            )

            for series_info in series_list:
                series_is_valid = True
                current_series_cells = set()

                # Check if any cell in the potential series is already excluded
                # Convert calamine indices to Excel coordinates
                for col_idx in range(series_info.start_col, series_info.end_col + 1):
                    excel_row = start_row + row_idx + 1
                    excel_col_letter = column_index_to_letter(start_col + col_idx)
                    cell_ref = f"{excel_col_letter}{excel_row}"
                    if cell_ref in exclude_cells:
                        series_is_valid = False
                        break
                    current_series_cells.add(cell_ref)

                if not series_is_valid:
                    continue

                # Create and add the new range - convert calamine indices to Excel coordinates
                start_excel_row = start_row + row_idx + 1
                start_excel_col_letter = column_index_to_letter(start_col + series_info.start_col)
                end_excel_col_letter = column_index_to_letter(start_col + series_info.end_col)
                start_cell = f"{start_excel_col_letter}{start_excel_row}"
                end_cell = f"{end_excel_col_letter}{start_excel_row}"
                
                time_series_range = FormatRange(
                    range=f"{start_cell}:{end_cell}",
                    type='date',
                    date_series_info={
                        'series_type': series_info.series_type,
                        'increment': series_info.increment,
                        'start_date': series_info.start_date,
                        'pattern': series_info.pattern
                    }
                )
                new_time_series_ranges.append(time_series_range)
                time_series_cells.update(current_series_cells)

        return new_time_series_ranges, time_series_cells

    def _detect_columnar_time_series(
        self,
        data: List[List],
        max_col: int,
        max_row: int,
        start_row: int,
        start_col: int,
        format_cache: dict,
        exclude_cells: set = None
    ) -> Tuple[List[FormatRange], set]:
        """
        Detect vertical (columnar) time series patterns using TimeRangeDetector.

        Args:
            data: Sheet data from calamine.
            max_col: Maximum column index.
            max_row: Maximum row index.
            start_row: Starting row offset from calamine (0-indexed).
            start_col: Starting column offset from calamine (0-indexed).
            format_cache: Dictionary mapping (row_idx, col_idx) to (number_format, format_id).
            exclude_cells: A set of cell references to exclude.

        Returns:
            A tuple of (new time series ranges, set of cells in these series).
        """
        if exclude_cells is None:
            exclude_cells = set()

        time_series_cells = set()
        new_time_series_ranges = []

        # Process column by column (transpose logic from row processing)
        for col_idx in range(max_col + 1):
            # Extract column data
            col_data = []
            col_formats = []

            for row_idx in range(max_row + 1):
                if row_idx < len(data):
                    row = data[row_idx]
                    cell_value = row[col_idx] if col_idx < len(row) else None
                else:
                    cell_value = None

                col_data.append(cell_value)

                # Get format code
                if (row_idx, col_idx) in format_cache:
                    number_format, *_ = format_cache[(row_idx, col_idx)]  # Unpack only number_format
                    col_formats.append(number_format)
                else:
                    col_formats.append(None)

            # Use detect_row_series for column data (it works the same way)
            # We pass col_idx as "row_number" since it represents the series identifier
            series_list = self.time_range_detector.detect_row_series(
                col_data, row_number=col_idx, start_col=0, format_codes=col_formats
            )

            # Only keep series with minimum length (columnar series should be longer)
            for series_info in series_list:
                series_length = series_info.end_col - series_info.start_col + 1
                if series_length < self.config.min_columnar_dates:
                    continue

                series_is_valid = True
                current_series_cells = set()

                # Check if any cell in the potential series is already excluded
                # For columns: series_info.start_col/end_col represent row indices
                for row_idx in range(series_info.start_col, series_info.end_col + 1):
                    excel_row = start_row + row_idx + 1
                    excel_col_letter = column_index_to_letter(start_col + col_idx)
                    cell_ref = f"{excel_col_letter}{excel_row}"
                    if cell_ref in exclude_cells:
                        series_is_valid = False
                        break
                    current_series_cells.add(cell_ref)

                if not series_is_valid:
                    continue

                # Create vertical range (same column, different rows)
                start_excel_row = start_row + series_info.start_col + 1
                end_excel_row = start_row + series_info.end_col + 1
                excel_col_letter = column_index_to_letter(start_col + col_idx)
                start_cell = f"{excel_col_letter}{start_excel_row}"
                end_cell = f"{excel_col_letter}{end_excel_row}"

                time_series_range = FormatRange(
                    range=f"{start_cell}:{end_cell}",
                    type='date',
                    date_series_info={
                        'series_type': series_info.series_type + '_column',  # Mark as column
                        'increment': series_info.increment,
                        'start_date': series_info.start_date,
                        'pattern': series_info.pattern
                    }
                )
                new_time_series_ranges.append(time_series_range)
                time_series_cells.update(current_series_cells)

        return new_time_series_ranges, time_series_cells

    def _detect_columnar_date_series(
        self,
        format_ranges: List[FormatRange],
        max_row: int
    ) -> Tuple[List[FormatRange], set]:
        """
        Detect vertical (columnar) date series and create consolidated ranges.

        Args:
            format_ranges: The list of format ranges from the initial scan.
            max_row: The maximum row number in the sheet.

        Returns:
            A tuple containing:
            - A list of new FormatRange objects for columnar date series.
            - A set of cell references (e.g., "C2") that are now part of a columnar series.
        """
        date_cells_by_column = {}
        
        # Step 1: Group all date cells by column
        for fmt_range in format_ranges:
            if fmt_range.type == 'date' and ':' not in fmt_range.range:
                col_idx = self._cell_to_col_index(fmt_range.range)
                row_idx = self._cell_to_row_index(fmt_range.range)
                
                if col_idx not in date_cells_by_column:
                    date_cells_by_column[col_idx] = []
                
                date_cells_by_column[col_idx].append({
                    "row": row_idx,
                    "date": fmt_range.parsed_date,
                    "cell_ref": fmt_range.range
                })

        new_columnar_ranges = []
        consolidated_cells = set()
        
        # Step 2: Analyze each column for unordered date series
        for col_idx, cells in date_cells_by_column.items():
            if len(cells) < self.config.min_columnar_dates:
                continue
                
            cells.sort(key=lambda x: x['row'])
            dates = [cell['date'] for cell in cells if cell['date']]

            if not dates:
                continue
            
            series_info = DateSeriesAnalyzer.analyze_date_range(dates)
            
            # Step 3: Create a new FormatRange if an unordered series is detected
            if series_info and series_info.get('interval_type') == 'unordered':
                min_row = cells[0]['row']
                max_row = cells[-1]['row']
                
                col_letter = column_index_to_letter(col_idx)
                start_cell = f"{col_letter}{min_row + 1}"
                end_cell = f"{col_letter}{max_row + 1}"
                
                column_range = FormatRange(
                    range=f"{start_cell}:{end_cell}",
                    type='date',
                    date_series_info={
                        'series_type': 'unordered_column',
                        'count': series_info.get('count'),
                        'start_date': series_info.get('first_date'),
                        'end_date': series_info.get('last_date')
                    }
                )
                new_columnar_ranges.append(column_range)
                
                # Step 4: Track all cells that are now part of this new range
                for cell in cells:
                    consolidated_cells.add(cell['cell_ref'])
                    
        return new_columnar_ranges, consolidated_cells

    def _cell_to_col_index(self, cell_ref: str) -> int:
        """Convert cell reference to 0-indexed column number (e.g., 'A1' -> 0, 'B1' -> 1)"""
        import re
        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref)
        if not match:
            return -1

        col_letters = match.group(1)
        col_idx = 0
        for char in col_letters:
            col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
        return col_idx - 1

    def _cell_to_row_index(self, cell_ref: str) -> int:
        """Convert cell reference to 0-indexed row number (e.g., 'A1' -> 0, 'A2' -> 1)"""
        import re
        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref)
        if not match:
            return -1

        return int(match.group(2)) - 1

    def _get_cells_in_range(self, range_str: str) -> set:
        """
        Get all cell references in a range.

        Args:
            range_str: Excel range like "G21:Q21" or "A1:C3"

        Returns:
            Set of cell references like {"G21", "H21", "I21", ...}
        """
        if ':' not in range_str:
            return {range_str}

        start_cell, end_cell = range_str.split(':')

        # Parse cells
        import re
        start_match = re.match(r'^([A-Z]+)(\d+)$', start_cell)
        end_match = re.match(r'^([A-Z]+)(\d+)$', end_cell)

        if not start_match or not end_match:
            return {range_str}

        start_col_str = start_match.group(1)
        start_row = int(start_match.group(2))
        end_col_str = end_match.group(1)
        end_row = int(end_match.group(2))

        # Convert column letters to numbers
        def col_to_num(col_str):
            num = 0
            for char in col_str:
                num = num * 26 + (ord(char) - ord('A') + 1)
            return num

        def num_to_col(n):
            result = ''
            while n > 0:
                n -= 1
                result = chr(65 + n % 26) + result
                n //= 26
            return result

        start_col_num = col_to_num(start_col_str)
        end_col_num = col_to_num(end_col_str)

        # Generate all cells in range
        cells = set()
        for row in range(start_row, end_row + 1):
            for col_num in range(start_col_num, end_col_num + 1):
                col_str = num_to_col(col_num)
                cells.add(f"{col_str}{row}")

        return cells

    def _get_row_from_range(self, range_str: str) -> int:
        """
        Extract row number from a range string for sorting.

        Args:
            range_str: Cell range like "A1", "A1:C1", or "AF2:DW2"

        Returns:
            Row number (1-indexed) from the start of the range
        """
        import re
        # Handle both single cells and ranges - just get the first cell
        first_cell = range_str.split(':')[0]

        # Extract row number (e.g., "AF2" -> 2)
        match = re.match(r'^([A-Z]+)(\d+)$', first_cell)
        if match:
            return int(match.group(2))

        # If parsing fails, return a large number to put it at the end
        return 999999

    def _get_col_from_range(self, range_str: str) -> int:
        """
        Extract column number from a range string for sorting.

        Args:
            range_str: Cell range like "A1", "A1:C1", or "AF2:DW2"

        Returns:
            Column number (0-indexed) from the start of the range
        """
        import re
        # Handle both single cells and ranges - just get the first cell
        first_cell = range_str.split(':')[0]

        # Extract column letters (e.g., "AF2" -> "AF")
        match = re.match(r'^([A-Z]+)(\d+)$', first_cell)
        if match:
            col_letters = match.group(1)
            # Convert column letters to number (A=0, B=1, ..., Z=25, AA=26, etc.)
            col_num = 0
            for char in col_letters:
                col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            return col_num - 1  # Return 0-indexed

        # If parsing fails, return a large number to put it at the end
        return 999999

    def _create_range(
        self,
        start_cell: str,
        end_cell: str,
        format_type: str,
        number_format: str,
        format_id: int,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        underline: Optional[bool] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None
    ) -> FormatRange:
        """
        Create a FormatRange object.

        Args:
            start_cell: Starting cell (e.g., "A1")
            end_cell: Ending cell (e.g., "C1")
            format_type: Format type classification
            number_format: Excel number format string
            format_id: Excel format ID
            bold: Optional bold styling
            italic: Optional italic styling
            underline: Optional underline styling
            font_color: Optional font/text color
            bg_color: Optional background color

        Returns:
            FormatRange object
        """
        if start_cell == end_cell:
            range_str = start_cell
        else:
            range_str = f"{start_cell}:{end_cell}"

        return FormatRange(
            range=range_str,
            type=format_type,
            format_code=number_format,
            format_id=format_id,
            bold=bold,
            italic=italic,
            underline=underline,
            font_color=font_color,
            bg_color=bg_color
        )
    
    def _merge_rectangular_regions(self, format_ranges: List[FormatRange]) -> List[FormatRange]:
        """
        Merge consecutive rows with identical column spans and format attributes
        into rectangular 2D regions.

        Args:
            format_ranges: Sorted list of format ranges

        Returns:
            New list with rectangular regions merged
        """
        if not format_ranges:
            return format_ranges

        # Build a dictionary of mergeable ranges by (row, start_col, end_col) -> index
        # This allows us to efficiently find consecutive rows without scanning the whole list
        mergeable_by_position = {}
        for i, fmt_range in enumerate(format_ranges):
            # Only consider ranges (not single cells) that are mergeable (not date/string)
            if ':' in fmt_range.range and fmt_range.type not in ('date', 'string'):
                row = self._get_row_from_range(fmt_range.range)
                start_col, _ = self._parse_cell_ref(fmt_range.range.split(':')[0])
                end_col, _ = self._parse_cell_ref(fmt_range.range.split(':')[1])
                key = (row, start_col, end_col)
                mergeable_by_position[key] = i

        # Track which ranges have been merged
        merged_indices = set()
        merged_ranges = []

        for i, current in enumerate(format_ranges):
            if i in merged_indices:
                continue

            # Skip single cells, dates, and strings (don't merge these)
            if ':' not in current.range or current.type in ('date', 'string'):
                merged_ranges.append(current)
                continue

            # Try to build a group of consecutive rows
            group = [current]
            merged_indices.add(i)

            # Get current range info
            curr_row = self._get_row_from_range(current.range)
            start_col, _ = self._parse_cell_ref(current.range.split(':')[0])
            end_col, _ = self._parse_cell_ref(current.range.split(':')[1])

            # Look for consecutive rows with same column span
            next_row = curr_row + 1
            while True:
                key = (next_row, start_col, end_col)
                if key not in mergeable_by_position:
                    break

                candidate_idx = mergeable_by_position[key]
                candidate = format_ranges[candidate_idx]

                # Check if formats match (including styling)
                if (current.type == candidate.type and
                    current.format_code == candidate.format_code and
                    current.bold == candidate.bold and
                    current.italic == candidate.italic and
                    current.underline == candidate.underline and
                    current.font_color == candidate.font_color and
                    current.bg_color == candidate.bg_color):
                    group.append(candidate)
                    merged_indices.add(candidate_idx)
                    next_row += 1
                else:
                    break

            # If we found multiple rows to merge, create a rectangular region
            if len(group) > 1:
                merged_range = self._create_rectangular_range(group)
                merged_ranges.append(merged_range)
            else:
                merged_ranges.append(current)

        return merged_ranges

    def _can_merge_into_rectangle(self, base_range: FormatRange, candidate: FormatRange) -> bool:
        """
        Check if a candidate range can be merged with a base range into a rectangle.

        Requirements:
        - Same format type
        - Same format attributes (scale, format_code)
        - Consecutive row numbers
        - Identical column span
        - Not a date or string type
        """
        # Must both be ranges (not single cells)
        if ':' not in candidate.range:
            return False

        # Skip dates and strings
        if candidate.type in ('date', 'string'):
            return False

        # Must have same type and format attributes (including styling)
        if (base_range.type != candidate.type or
            base_range.format_code != candidate.format_code or
            base_range.bold != candidate.bold or
            base_range.italic != candidate.italic or
            base_range.underline != candidate.underline or
            base_range.font_color != candidate.font_color or
            base_range.bg_color != candidate.bg_color):
            return False

        # Parse ranges
        base_start, base_end = base_range.range.split(':')
        cand_start, cand_end = candidate.range.split(':')

        # Extract columns and rows
        base_start_col, base_start_row = self._parse_cell_ref(base_start)
        base_end_col, base_end_row = self._parse_cell_ref(base_end)
        cand_start_col, cand_start_row = self._parse_cell_ref(cand_start)
        cand_end_col, cand_end_row = self._parse_cell_ref(cand_end)

        # Must have identical column spans
        if (base_start_col != cand_start_col or base_end_col != cand_end_col):
            return False

        # Must be consecutive rows (candidate should be next row after base)
        if cand_start_row != base_end_row + 1:
            return False

        return True

    def _create_rectangular_range(self, group: List[FormatRange]) -> FormatRange:
        """
        Create a single FormatRange representing a rectangular region from a group
        of consecutive row ranges.
        """
        first = group[0]
        last = group[-1]

        # Parse first and last ranges
        first_start, _ = first.range.split(':')
        _, last_end = last.range.split(':')

        # Create new rectangular range
        return FormatRange(
            range=f"{first_start}:{last_end}",
            type=first.type,
            format_code=first.format_code,
            format_id=first.format_id,
            bold=first.bold,
            italic=first.italic,
            underline=first.underline,
            font_color=first.font_color,
            bg_color=first.bg_color
        )

    def _parse_cell_ref(self, cell_ref: str) -> Tuple[str, int]:
        """
        Parse a cell reference into column letter(s) and row number.

        Args:
            cell_ref: Cell reference like "E8" or "AB10"

        Returns:
            Tuple of (column_letters, row_number)
        """
        import re
        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref)
        if not match:
            return ("", 0)
        return (match.group(1), int(match.group(2)))

    def _get_first_sheet_name(self, file_path: str) -> str:
        """Get the name of the first sheet in the workbook"""
        try:
            from python_calamine import CalamineWorkbook
            workbook = CalamineWorkbook.from_path(file_path)
            sheet_names = workbook.sheet_names
            return sheet_names[0] if sheet_names else "Sheet1"
        except Exception:
            return "Sheet1"  # Fallback