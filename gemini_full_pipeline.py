"""
Full Gemini pipeline: boundary detection → compression → LLM analysis
Processes all sheets in a workbook end-to-end.
"""
import json
import time
import sys
import argparse
import re
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Tuple, Optional, List

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("Note: tqdm not installed. Install with 'pip install tqdm' for progress bars.")

try:
    import vertexai
    from vertexai.generative_models import GenerativeModel
    VERTEXAI_AVAILABLE = True
except ImportError:
    VERTEXAI_AVAILABLE = False
    print("ERROR: google-cloud-aiplatform not installed.")
    print("Install with: pip install google-cloud-aiplatform")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Sheet ordering may be alphabetical.")
    print("Install with: pip install openpyxl")

# Import boundary detection and compression modules
from boundary_detector.core.batch_detector import BatchBoundaryDetector
from compression.format_compression import compress_formats
from compression.inverted_index_compression import compress_format_result_json, save_compressed_json

# Base paths
BASE_DIR = Path(__file__).parent
SERVICE_ACCOUNT_FILE = BASE_DIR / "gemini_service_account.json"

# Gemini prompt with label/data separation and multiple time series detection
ANALYSIS_PROMPT = """You are an expert financial analyst tasked with creating a high-level summary of an Excel spreadsheet's structure. Your goal is to identify and describe the key logical data sections (e.g., "Income Statement," "Balance Sheet," "KPI Dashboard") without listing every single row or label. The output must be a structured markdown document that enables retrieval of specific, relevant sections by cell range.

# CORE OBJECTIVE
Identify major financial or data sections within this Excel sheet, describe their purpose, and define their precise boundaries and time series context.

# CRITICAL: MULTIPLE TIME SERIES DETECTION
Many spreadsheets have MULTIPLE time series sections side-by-side:
- **Annual labels** (e.g., rows or columns showing years like 2015, 2016, 2017...)
- **Monthly/Quarterly labels** (e.g., rows or columns showing monthly dates like 2015-01-31, 2015-02-28... or quarters like Q1 2024, Q2 2024...)

When analyzing time series:
1. Look for MULTIPLE date header rows or columns with different frequencies
2. Check if there are gaps between column groups (suggesting separate time series)
3. Identify if data spans different date column groups
4. Report ALL time series horizons found, not just one

# OUTPUT REQUIREMENTS
Generate a markdown document with the following EXACT structure for each sheet:

## 1. **Sheet Name**: The name of the analyzed sheet.

For EACH logical section identified, create a subsection.

### Section Name (e.g., "Income Statement")
- **Section Type**: Classify the section. Examples: `Standard P&L`, `Custom P&L`, `Balance Sheet`, `Waterfall Chart Data`, `Key Metrics Table`.
- **Description & Purpose**: Briefly explain what this section represents and its business purpose.
- **Cell Range**: The exact and complete cell range for this entire logical section (e.g., `A5:FS88`). This is critical for future data retrieval.
- **Layout Structure**:
    - **Row Headers Location**
    - **Column Headers Location**
    - **Data Range**: The cell range containing the actual numeric values/formulas. If there are multiple time series, specify each:
      - Annual data: `E7:Q24` (numeric values for annual periods)
      - Monthly data: `T7:FS24` (numeric values for monthly periods)
- **Time Series Details**:
    - **Date Range**: The start and end dates/periods covered for EACH time series:
      - Annual: 2015 to 2027 (example)
      - Monthly: 2015-01-31 to 2027-12-31 (example)
    - **Frequency**: The time interval for EACH series (e.g., `Annual`, `Monthly`, `Quarterly`)
- **Key Components**: List the major row labels/metrics in this section (e.g., Revenue, COGS, Gross Profit, EBITDA). DO NOT list every single row.
- **Notes & Customizations**: Mention any deviations from standard reports, or special calculations.

# GUIDING PRINCIPLES
- **High-Level, Not Granular**: Focus on the forest, not the trees. Summarize sections, don't transcribe them.
- **Precision is Key**: Cell ranges and date locations must be exact for future retrieval.
- **Context Over Content**: Explain the *purpose* of a section, not just what's in it.
- **Multiple Time Series**: If you see evidence of multiple date columns with different frequencies, report them ALL.
- **Date Anchor Points**: If the JSON contains a `date_series_definitions` section with `anchor_points`, include these in your Time Series Details to help locate specific dates within long series (e.g., "Monthly: 2015-01-31 to 2027-12-31 (T3:FS3). Anchor points: T3=2015-01-31, AF3=2016-01-31, AR3=2017-01-31...").

Now, analyze the provided compressed JSON and generate the complete, high-level structured documentation."""


def find_excel_files(raw_data_dir: Path) -> List[Path]:
    """Find all Excel files in a directory"""
    excel_files = []
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(raw_data_dir.glob(ext))
    return sorted(excel_files)


def get_dataset_paths(dataset_name: str) -> Tuple[Path, Path, Path]:
    """
    Get paths for a dataset

    Returns:
        Tuple of (raw_data_dir, processing_outputs_dir, dataset_base_dir)
    """
    dataset_base = BASE_DIR / "datasets" / dataset_name
    raw_data_dir = dataset_base / "raw_data"
    processing_outputs_dir = dataset_base / "processing_outputs"

    return raw_data_dir, processing_outputs_dir, dataset_base


def select_excel_file(excel_files: List[Path]) -> Optional[Path]:
    """
    Prompt user to select an Excel file if multiple are found

    Returns:
        Selected file path or None if cancelled
    """
    if len(excel_files) == 0:
        return None

    if len(excel_files) == 1:
        return excel_files[0]

    # Multiple files - prompt user
    print("\nMultiple Excel files found:")
    for i, file_path in enumerate(excel_files, 1):
        print(f"  {i}. {file_path.name}")

    while True:
        try:
            choice = input(f"\nSelect a file (1-{len(excel_files)}) or 'q' to quit: ").strip()
            if choice.lower() == 'q':
                return None

            idx = int(choice) - 1
            if 0 <= idx < len(excel_files):
                return excel_files[idx]
            else:
                print(f"Invalid choice. Please enter a number between 1 and {len(excel_files)}.")
        except ValueError:
            print("Invalid input. Please enter a number or 'q' to quit.")
        except (EOFError, KeyboardInterrupt):
            print("\nCancelled.")
            return None


def sanitize_filename(name: str) -> str:
    """Convert sheet name to safe filename"""
    safe = name.replace('/', '_').replace('\\', '_').replace(':', '_')
    safe = safe.replace('*', '_').replace('?', '_').replace('"', '_')
    safe = safe.replace('<', '_').replace('>', '_').replace('|', '_')
    return safe


def get_workbook_sheet_order(workbook_path: Path) -> List[str]:
    """
    Get the sheet names in the order they appear in the workbook.

    Returns:
        List of sheet names in workbook order
    """
    if not OPENPYXL_AVAILABLE:
        return []

    try:
        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"WARNING: Could not load workbook to determine sheet order: {e}")
        print("Falling back to alphabetical sorting")
        return []


def parse_markdown_file(file_path: Path) -> Dict:
    """
    Parse a markdown file to extract sheet name and sections.

    Returns:
        Dict with 'sheet_name', 'sections', and 'content'
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Remove markdown code fence if present
    content = re.sub(r'^```markdown\n', '', content)
    content = re.sub(r'\n```$', '', content)

    # Extract sheet name from "## 1. Spreadsheet Overview" section
    sheet_name_match = re.search(r'\*\*Sheet Name\*\*:\s*(.+?)(?:\n|$)', content)
    sheet_name = sheet_name_match.group(1).strip() if sheet_name_match else file_path.stem.replace('_analysis', '')

    # Extract all section headers (### lines)
    sections = []
    for match in re.finditer(r'^### (.+?)$', content, re.MULTILINE):
        section_title = match.group(1).strip()
        sections.append(section_title)

    return {
        'sheet_name': sheet_name,
        'file_path': file_path,
        'sections': sections,
        'content': content
    }


def create_master_toc(parsed_files: List[Dict]) -> str:
    """
    Create a master table of contents for all sheets.

    Returns:
        Markdown string with TOC
    """
    toc = ["# Master Workbook Analysis", ""]
    toc.append("This document provides a comprehensive analysis of all sheets in the Excel workbook.")
    toc.append("")
    toc.append("---")
    toc.append("")
    toc.append("## Table of Contents")
    toc.append("")

    for idx, parsed in enumerate(parsed_files, 1):
        sheet_name = parsed['sheet_name']
        # Create anchor link (lowercase, replace spaces with hyphens)
        anchor = sheet_name.lower().replace(' ', '-').replace('/', '').replace('\\', '').replace('-->', '').replace('>', '').replace('<', '').replace('|', '').replace(':', '')
        toc.append(f"{idx}. [{sheet_name}](#{anchor})")

        # Add subsections if any
        if parsed['sections']:
            for section in parsed['sections']:
                # Create section anchor
                section_anchor = section.lower().replace(' ', '-').replace('/', '').replace('\\', '').replace('(', '').replace(')', '').replace('.', '').replace(',', '').replace(':', '')
                toc.append(f"   - [{section}](#{anchor}-{section_anchor})")

    toc.append("")
    toc.append("---")
    toc.append("")

    return '\n'.join(toc)


def format_sheet_section(parsed: Dict, sheet_number: int) -> str:
    """
    Format a single sheet's content with clear differentiation.

    Returns:
        Markdown string for the sheet
    """
    sheet_name = parsed['sheet_name']
    content = parsed['content']

    # Create a clear sheet header
    output = []
    output.append("")
    output.append("=" * 100)
    output.append(f"# SHEET {sheet_number}: {sheet_name}")
    output.append("=" * 100)
    output.append("")

    # Add the sheet content (remove the first # if it exists to avoid nesting issues)
    content = re.sub(r'^# ', '## ', content, flags=re.MULTILINE)

    output.append(content)
    output.append("")

    return '\n'.join(output)


def combine_markdown_files(workbook_path: Path, markdown_dir: Path, output_file: Path) -> bool:
    """
    Combine all markdown files into a single master document.

    Args:
        workbook_path: Path to the Excel workbook
        markdown_dir: Directory containing individual sheet markdown files
        output_file: Path for the combined output file

    Returns:
        True if successful, False otherwise
    """
    print("\nSTEP 3: Combining markdown analyses into master document...")

    # Get workbook sheet order
    workbook_sheet_order = get_workbook_sheet_order(workbook_path)
    if workbook_sheet_order:
        print(f"  Found {len(workbook_sheet_order)} sheets in workbook order")

    # Find all markdown files and create a lookup by sheet name
    markdown_files_by_sheet = {}
    for file_path in markdown_dir.glob("*_analysis.md"):
        # Extract original sheet name from filename
        safe_name = file_path.stem.replace('_analysis', '')
        # Try to match to original sheet name
        for sheet_name in workbook_sheet_order:
            if sanitize_filename(sheet_name) == safe_name:
                markdown_files_by_sheet[sheet_name] = file_path
                break
        else:
            # If no match found, store with sanitized name as fallback
            markdown_files_by_sheet[safe_name] = file_path

    if not markdown_files_by_sheet:
        print("  ERROR: No markdown files found to combine")
        return False

    print(f"  Found {len(markdown_files_by_sheet)} markdown files")

    # Parse all files in workbook order
    parsed_files = []

    # First, add sheets in workbook order if we have the sheet order
    if workbook_sheet_order:
        for sheet_name in workbook_sheet_order:
            if sheet_name in markdown_files_by_sheet:
                file_path = markdown_files_by_sheet[sheet_name]
                try:
                    parsed = parse_markdown_file(file_path)
                    parsed_files.append(parsed)
                except Exception as e:
                    print(f"  WARNING: Could not parse {file_path.name}: {e}")
    else:
        # Fallback to alphabetical if we couldn't get workbook order
        print("  Using alphabetical order...")
        for sheet_name in sorted(markdown_files_by_sheet.keys()):
            file_path = markdown_files_by_sheet[sheet_name]
            try:
                parsed = parse_markdown_file(file_path)
                parsed_files.append(parsed)
            except Exception as e:
                print(f"  WARNING: Could not parse {file_path.name}: {e}")

    if not parsed_files:
        print("  ERROR: No markdown files could be parsed")
        return False

    # Create master document
    master_content = []

    # Add table of contents
    master_content.append(create_master_toc(parsed_files))

    # Add each sheet
    for idx, parsed in enumerate(parsed_files, 1):
        sheet_section = format_sheet_section(parsed, idx)
        master_content.append(sheet_section)

    # Write to file
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(master_content))

    print(f"  Combined {len(parsed_files)} sheets into master document")
    print(f"  Output: {output_file}")
    print(f"  File size: {output_file.stat().st_size:,} bytes")

    return True


def process_sheet(
    sheet_name: str,
    workbook_path: Path,
    boundary_result,
    project_id: str,
    json_dir: Path,
    markdown_dir: Path,
    max_retries: int = 5
) -> Tuple[str, bool, Optional[str]]:
    """
    Process a single sheet: compress to JSON → analyze with Gemini

    Args:
        sheet_name: Name of the sheet to process
        workbook_path: Path to Excel workbook
        boundary_result: BoundaryResult object from boundary detection
        project_id: Google Cloud project ID
        json_dir: Directory to save JSON output
        markdown_dir: Directory to save markdown output
        max_retries: Max retry attempts for rate limits

    Returns:
        Tuple of (sheet_name, success, error_message)
    """
    if not VERTEXAI_AVAILABLE:
        return (sheet_name, False, "vertexai SDK not installed")

    safe_name = sanitize_filename(sheet_name)
    json_file = json_dir / f"{safe_name}_compressed.json"
    markdown_file = markdown_dir / f"{safe_name}_analysis.md"

    # Skip if markdown already exists
    if markdown_file.exists():
        return (sheet_name, True, None)

    try:
        # Step 1: Compress sheet to JSON (if not already done)
        if not json_file.exists():
            # Compress formats using the boundary result
            format_result = compress_formats(
                file_path=str(workbook_path),
                sheet_name=sheet_name,
                boundary_result=boundary_result
            )

            # Convert FormatResult to JSON
            json_output = format_result.to_json()

            # Apply inverted index compression
            compressed_json = compress_format_result_json(json_output)

            # Save compressed JSON
            json_dir.mkdir(parents=True, exist_ok=True)
            save_compressed_json(compressed_json, str(json_file))

        # Step 2: Load compressed JSON
        with open(json_file, 'r', encoding='utf-8') as f:
            compressed_json = json.load(f)

        # Step 3: Analyze with Gemini (with retry logic)
        for attempt in range(max_retries):
            try:
                # Convert JSON to string for prompt
                json_str = json.dumps(compressed_json, indent=2)

                # Create the full prompt
                full_prompt = f"{ANALYSIS_PROMPT}\n\n# Compressed JSON Data\n\n```json\n{json_str}\n```"

                # Initialize Vertex AI
                vertexai.init(project=project_id, location="us-central1")

                # Create model
                model = GenerativeModel("gemini-2.0-flash-exp")

                # Add a small delay to avoid overwhelming the API (spread out requests)
                if attempt > 0:
                    time.sleep(2)  # Wait 2s before retry

                # Generate response
                response = model.generate_content(
                    full_prompt,
                    generation_config={
                        "max_output_tokens": 8192,
                        "temperature": 0.3,
                    }
                )

                # Extract text
                analysis = response.text

                if not analysis or len(analysis.strip()) == 0:
                    return (sheet_name, False, "Empty response from Gemini")

                # Save markdown
                markdown_dir.mkdir(parents=True, exist_ok=True)
                with open(markdown_file, 'w', encoding='utf-8') as f:
                    f.write(analysis)

                return (sheet_name, True, None)

            except Exception as e:
                error_msg = str(e)

                # Check if it's a rate limit error (multiple possible error messages)
                is_rate_limit = (
                    "429" in error_msg or
                    "quota" in error_msg.lower() or
                    "rate" in error_msg.lower() or
                    "resource_exhausted" in error_msg.lower() or
                    "RESOURCE_EXHAUSTED" in error_msg
                )

                if is_rate_limit:
                    if attempt < max_retries - 1:
                        # Exponential backoff: 10s, 20s, 40s, 80s
                        wait_time = (2 ** attempt) * 10
                        print(f"\n  ⏸️  Rate limit hit for '{sheet_name}', waiting {wait_time}s... (attempt {attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        # Max retries exceeded for rate limit
                        return (sheet_name, False, f"Rate limit - max retries exceeded: {error_msg[:100]}")

                # Non-rate-limit error - fail immediately
                return (sheet_name, False, error_msg[:200])

        return (sheet_name, False, "Max retries exceeded")

    except Exception as e:
        return (sheet_name, False, str(e))


def main():
    """Run full Gemini pipeline on entire workbook"""

    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="Run Gemini pipeline on a dataset",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python gemini_full_pipeline.py alphasense
  python gemini_full_pipeline.py evergreen-asset-acquisition
  python gemini_full_pipeline.py pendo-sales-quota-build
        """
    )
    parser.add_argument(
        "dataset",
        help="Name of the dataset folder (e.g., 'alphasense', 'evergreen-asset-acquisition')"
    )
    args = parser.parse_args()

    dataset_name = args.dataset

    # Get dataset paths
    raw_data_dir, processing_outputs_dir, dataset_base = get_dataset_paths(dataset_name)

    # Verify dataset exists
    if not dataset_base.exists():
        print(f"ERROR: Dataset folder not found: {dataset_base}")
        print(f"\nAvailable datasets:")
        datasets_dir = BASE_DIR / "datasets"
        if datasets_dir.exists():
            for item in sorted(datasets_dir.iterdir()):
                if item.is_dir():
                    print(f"  - {item.name}")
        return

    if not raw_data_dir.exists():
        print(f"ERROR: raw_data folder not found: {raw_data_dir}")
        return

    # Find Excel files
    excel_files = find_excel_files(raw_data_dir)
    if not excel_files:
        print(f"ERROR: No Excel files found in {raw_data_dir}")
        return

    # Select Excel file
    workbook_path = select_excel_file(excel_files)
    if workbook_path is None:
        print("No file selected. Exiting.")
        return

    # Setup output directories
    json_dir = processing_outputs_dir / "json"
    markdown_dir = processing_outputs_dir / "markdown"

    print(f"\n{'='*80}")
    print(f"GEMINI FULL PIPELINE")
    print(f"{'='*80}")
    print(f"Dataset: {dataset_name}")
    print(f"Workbook: {workbook_path}")
    print(f"JSON Output: {json_dir}")
    print(f"Markdown Output: {markdown_dir}")
    print(f"{'='*80}\n")

    if not VERTEXAI_AVAILABLE:
        print("\nERROR: Cannot proceed without google-cloud-aiplatform")
        print("Install with: pip install google-cloud-aiplatform")
        return

    # Set service account environment variable
    import os
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(SERVICE_ACCOUNT_FILE)

    # Load project ID from service account
    if not SERVICE_ACCOUNT_FILE.exists():
        print(f"ERROR: Service account file not found: {SERVICE_ACCOUNT_FILE}")
        return

    with open(SERVICE_ACCOUNT_FILE, 'r') as f:
        sa_data = json.load(f)
    project_id = sa_data['project_id']

    # Step 1: Detect boundaries for all sheets
    print("STEP 1: Detecting boundaries for all sheets...")
    start_time = time.time()

    detector = BatchBoundaryDetector()
    boundary_results = detector.detect_all_sheets(
        str(workbook_path),
        detect_charts=False,  # Disable chart detection for speed (saves ~30s)
        parallel=True,
        max_workers=5
    )

    elapsed = time.time() - start_time
    print(f"  Boundary detection complete: {elapsed:.1f}s")
    print(f"  Found {len(boundary_results)} sheets\n")

    # Step 2: Process all sheets (compression + LLM analysis) in parallel
    print("STEP 2: Compressing sheets and analyzing with Gemini...")
    print("  Using 3 parallel workers to respect API rate limits")
    start_time = time.time()
    success_count = 0
    error_count = 0
    errors = []

    # Use 3 workers to balance speed with API rate limits
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {
            executor.submit(
                process_sheet,
                sheet_name,
                workbook_path,
                boundary_result,
                project_id,
                json_dir,
                markdown_dir
            ): sheet_name
            for sheet_name, boundary_result in boundary_results.items()
        }

        # Use tqdm if available
        if TQDM_AVAILABLE:
            iterator = tqdm(as_completed(futures), total=len(futures), desc="Processing")
        else:
            iterator = as_completed(futures)

        # Collect results
        for future in iterator:
            sheet_name, success, error = future.result()

            if success:
                success_count += 1
                if not TQDM_AVAILABLE:
                    print(f"  OK {sheet_name}")
            else:
                error_count += 1
                errors.append((sheet_name, error))
                if not TQDM_AVAILABLE:
                    print(f"  FAIL {sheet_name}: {error}")

    elapsed = time.time() - start_time
    print(f"\n  Compression and analysis complete: {elapsed:.1f}s ({elapsed/60:.1f} minutes)")

    # Step 3: Combine all markdown files into master document
    master_output_file = processing_outputs_dir / "MASTER_WORKBOOK_ANALYSIS.md"
    combine_success = combine_markdown_files(workbook_path, markdown_dir, master_output_file)

    # Summary
    print(f"\n{'='*80}")
    print(f"PIPELINE COMPLETE")
    print(f"{'='*80}")
    print(f"Total sheets: {len(boundary_results)}")
    print(f"Successful: {success_count}/{len(boundary_results)}")
    print(f"Failed: {error_count}/{len(boundary_results)}")
    if errors:
        print(f"\nFailed sheets:")
        for sheet_name, error in errors:
            print(f"  - {sheet_name}: {error}")
    print(f"\nProcessing time: {elapsed:.1f}s ({elapsed/60:.1f} minutes)")
    print(f"\nOutputs saved to:")
    print(f"  JSON: {json_dir}")
    print(f"  Markdown (individual sheets): {markdown_dir}")
    if combine_success:
        print(f"  Master document: {master_output_file}")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()
